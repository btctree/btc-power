"""Full report on the ORIGINAL 8B model (5x, NO turnover control) + VOL+FUND:
 1) Sizing/Cap grid (legacy flat-lev vs size-control cap x vol-target) @0/50/100bp + metrics + margin use
 2) Margin-level sweep (maintenance-margin / liquidation threshold) -> liquidation sensitivity
 3) Equity curves + charts. Honest: 2014+, VOL+FUND gates, dd-kill, honest liquidation.
"""
import os
import numpy as np, pandas as pd
import stable_combo as sc
import live_engine as le
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
ANN = 365
OUT = r"C:\Users\user\OneDrive\Desktop\New setup for BTC\BTC_8B_sizing_margin_report.xlsx"
df, reg0, memb = sc.prep()
reg, emap, exp_raw = le.ensemble_ctx(df, memb)
expf = np.where(np.abs(exp_raw) >= 0.4, exp_raw, 0.0)
close = df["close"].to_numpy(); low = df["low"].to_numpy(); high = df["high"].to_numpy()
n = len(df); i0 = 260; dstr = df["Date"].tolist(); dates = pd.to_datetime(df["Date"])
rv = pd.Series(close).pct_change().rolling(20).std().to_numpy() * np.sqrt(ANN)
HERE = os.path.dirname(os.path.abspath(__file__))
def lmap(p, col):
    f = pd.read_csv(p); return dict(zip(f.iloc[:, 0].astype(str), f[col]))
funding = np.array([lmap(os.path.join(HERE, "..", "data", "funding.csv"), "funding_rate").get(d, np.nan) for d in dstr])
def trk(a, w=365): return pd.Series(a).rolling(w, min_periods=60).apply(lambda x: (x.iloc[-1] >= x).mean(), raw=False).to_numpy()
vr = trk(rv); fr = trk(funding)
gl = np.ones(n); gs = np.ones(n)
for i in range(n):
    if vr[i] == vr[i] and vr[i] > 0.85: gl[i] *= 0.5; gs[i] *= 0.5
    if fr[i] == fr[i]:
        if fr[i] > 0.90: gl[i] *= 0.5
        if fr[i] < 0.10: gs[i] *= 0.5
WIN = [("2017-12-16","2018-11-18"),("2021-10-20","2022-03-09"),("2025-05-22","2025-12-01")]

def sim(slip, legacy=None, cap=5.0, vt=0.6, acct_lev=5.0, dd_kill=0.30, maint=0.01, fee=0.0005):
    # ORIGINAL 8B = NO turnover control (no smoothing, no deadband). VOL+FUND gates on.
    e_in = expf
    eqv = peak = 500.0; held = 0.0; eq = np.full(n, 500.0); E = np.zeros(n); liq = 0; margs = []
    for i in range(i0, n):
        sig = e_in[i - 1]; g = gl[i - 1] if sig > 0 else (gs[i - 1] if sig < 0 else 1.0)
        if not (rv[i - 1] == rv[i - 1] and rv[i - 1] > 0): eq[i] = eqv; E[i] = held; continue
        if legacy: mult = legacy * min(1.0, vt / rv[i - 1])      # original-style flat leverage (scale-down only)
        else: mult = min(cap, vt / rv[i - 1])                    # size-control (scale up in calm to cap)
        e = sig * g * mult
        if dd_kill > 0 and eqv < peak * (1 - dd_kill): e *= 0.5
        adv = (-(low[i] / close[i - 1] - 1)) if e > 0 else ((high[i] / close[i - 1] - 1) if e < 0 else 0.0)
        if e != 0 and abs(e) * max(adv, 0) >= (1 - maint):
            eqv *= 0.01; liq += 1; held = 0.0; eq[i] = eqv; E[i] = 0.0; peak = max(peak, eqv); continue
        eqv *= (1 + e * (close[i] / close[i - 1] - 1)); eqv -= eqv * abs(e - held) * (fee + slip)
        held = e; eq[i] = max(eqv, 1e-9); E[i] = e; peak = max(peak, eqv); margs.append(abs(e) / acct_lev)
    return eq, E, liq, (np.mean(margs) if margs else 0), (np.max(margs) if margs else 0)
def metr(eq, E):
    s = pd.Series(eq[i0:], index=dates.iloc[i0:].values); r = s.pct_change().dropna(); yrs = len(s) / ANN
    cagr = (s.iloc[-1] / 500) ** (1 / yrs) - 1 if s.iloc[-1] > 0 else -1
    sh = r.mean() * ANN / (r.std(ddof=1) * np.sqrt(ANN)) if r.std() > 0 else float("nan")
    dd = (s / s.cummax() - 1).min(); ww = [(s.loc[a:b]/s.loc[a:b].cummax()-1).min()*100 for a,b in WIN]
    i = i0; w = 0; tot = 0
    while i < n:
        sg = 1 if E[i] > 0 else (-1 if E[i] < 0 else 0)
        if sg == 0: i += 1; continue
        j = i
        while j+1 < n and (1 if E[j+1]>0 else (-1 if E[j+1]<0 else 0)) == sg: j += 1
        f0 = eq[i-1] if i > 0 else 500.0
        if eq[j]/f0-1 > 0: w += 1
        tot += 1; i = j + 1
    return s.iloc[-1], cagr, sh, (cagr/abs(dd) if dd<0 else 0), dd, (w/tot if tot else 0), tot, ww, s

# ---- (1) sizing/cap configs ----
SC = [("Legacy 2x (flat)", dict(legacy=2, acct_lev=2)),
      ("Legacy 3x (flat)", dict(legacy=3, acct_lev=3)),
      ("Legacy 5x = ORIGINAL 8B", dict(legacy=5, acct_lev=5)),
      ("SizeCtrl cap2 vt0.6", dict(cap=2, vt=0.6, acct_lev=2)),
      ("SizeCtrl cap3 vt0.6", dict(cap=3, vt=0.6, acct_lev=3)),
      ("SizeCtrl cap3 vt0.8", dict(cap=3, vt=0.8, acct_lev=3)),
      ("SizeCtrl cap5 vt0.6", dict(cap=5, vt=0.6, acct_lev=5)),
      ("SizeCtrl cap5 vt0.8", dict(cap=5, vt=0.8, acct_lev=5)),
      ("SizeCtrl cap5 vt1.5", dict(cap=5, vt=1.5, acct_lev=5))]
rows = []
print(f"{'config':26s} {'$@0bp':>13s} {'$@50bp':>12s} {'$@100bp':>11s} {'CAGR':>5s} {'Calm':>5s} {'Shrp':>5s} {'maxDD':>6s} {'Win%':>5s} {'liq':>4s} {'avgMrg':>6s} {'maxMrg':>6s}")
for nm, kw in SC:
    eq0,E0,l0,am0,mm0 = sim(0.0, **kw); eq5,E5,l5,am,mm = sim(0.005, **kw); eq1,E1,l1,_,_ = sim(0.010, **kw)
    f0=eq0[-1]; f5,cagr,sh,cal,dd,wr,tot,ww,s5 = metr(eq5,E5); f1=eq1[-1]
    rows.append((nm,f0,f5,f1,cagr,cal,sh,dd,wr,l5,am,mm,ww,tot,s5))
    print(f"{nm:26s} ${f0:>12,.0f} ${f5:>11,.0f} ${f1:>10,.0f} {cagr*100:>4.0f}% {cal:>5.2f} {sh:>5.2f} {dd*100:>5.0f}% {wr*100:>4.0f}% {l5:>4d} {am*100:>5.0f}% {mm*100:>5.0f}%")

# ---- (2) margin-level (maintenance / liquidation threshold) sweep ----
print("\n=== MARGIN LEVEL sweep (maintenance margin = how big an adverse move liquidates) ===")
marg_rows = []
for base_nm, kw in [("ORIGINAL 8B 5x", dict(legacy=5, acct_lev=5)), ("SizeCtrl cap5 vt0.6", dict(cap=5, vt=0.6, acct_lev=5)), ("SizeCtrl cap3 vt0.6", dict(cap=3, vt=0.6, acct_lev=3))]:
    for maint in [0.005, 0.01, 0.025, 0.05, 0.10]:
        eq,E,liq,am,mm = sim(0.005, maint=maint, **kw); f5,cagr,sh,cal,dd,wr,tot,ww,s5 = metr(eq,E)
        marg_rows.append((base_nm, maint, f5, dd, liq))
        print(f"  {base_nm:20s} maint {maint*100:>4.1f}% (liq at ~{(1-maint)/kw.get('acct_lev',5)*100:.0f}% move) -> $50bp {f5:>11,.0f} | maxDD {dd*100:>4.0f}% | liq {liq}")

# ================= Excel =================
HDR=Font(bold=True,color="FFFFFF"); HF=PatternFill("solid",fgColor="7F1D1D"); TIT=Font(bold=True,size=13)
thin=Side(style="thin",color="DDDDDD"); BD=Border(thin,thin,thin,thin)
wb=Workbook(); ws=wb.active; ws.title="Sizing & Cap"
ws["A1"]="ORIGINAL 8B (5x, NO turnover control) + VOL+FUND — sizing/cap report"; ws["A1"].font=TIT
ws["A2"]="$500 from 2014. Legacy=flat leverage (scale-down only); SizeCtrl=min(cap, vt/rv) scales up in calm. Margin% = exposure/account-leverage."; ws["A2"].font=Font(italic=True,color="666666")
cols=["Config","$@0bp","$@50bp","$@100bp","CAGR","Calmar","Sharpe","maxDD","Win%","Liq","avg Margin%","max Margin%","W1","W2","W3"]
for j,c in enumerate(cols,1):
    cc=ws.cell(4,j,c); cc.font=HDR; cc.fill=HF; cc.alignment=Alignment(horizontal="center",wrap_text=True)
for k,(nm,f0,f5,f1,cagr,cal,sh,dd,wr,liq,am,mm,ww,tot,s5) in enumerate(rows):
    r=5+k; vals=[nm,round(f0),round(f5),round(f1),cagr,cal,sh,dd,wr,liq,am,mm,ww[0]/100,ww[1]/100,ww[2]/100]
    for j,v in enumerate(vals,1):
        cc=ws.cell(r,j,v); cc.border=BD
    for j in (2,3,4): ws.cell(r,j).number_format='#,##0'
    for j in (5,9,11,12,13,14,15): ws.cell(r,j).number_format='0%'
    for j in (6,7): ws.cell(r,j).number_format='0.00'
    ws.cell(r,8).number_format='0%'
    if "ORIGINAL" in nm:
        for j in range(1,16): ws.cell(r,j).fill=PatternFill("solid",fgColor="FDE2E2")
ws.column_dimensions["A"].width=26
for c in "BCDEFGHIJKLMNO": ws.column_dimensions[c].width=11

wm=wb.create_sheet("Margin Levels")
wm["A1"]="Margin-level (maintenance-margin) sweep — how the liquidation threshold affects results @50bp"; wm["A1"].font=TIT
mh=["Base model","Maint margin","Liq move ≈","$@50bp","maxDD","Liquidations"]
for j,h in enumerate(mh,1):
    cc=wm.cell(3,j,h); cc.font=HDR; cc.fill=HF
for k,(bn,maint,f5,dd,liq) in enumerate(marg_rows):
    r=4+k; al=5 if "cap3" not in bn else 3
    for j,v in enumerate([bn, maint, (1-maint)/al, round(f5), dd, liq],1):
        wm.cell(r,j,v)
    wm.cell(r,2).number_format='0.0%'; wm.cell(r,3).number_format='0%'; wm.cell(r,4).number_format='#,##0'; wm.cell(r,5).number_format='0%'
wm.column_dimensions["A"].width=20; wm.column_dimensions["B"].width=12; wm.column_dimensions["C"].width=10; wm.column_dimensions["D"].width=14

we=wb.create_sheet("Equity Curves")
chart_idx=[0,2,4,6]  # legacy2, orig8B5x, sizectrl cap3, sizectrl cap5 vt0.6
series=[(rows[i][0], rows[i][14]) for i in chart_idx]
ddt=[d.strftime("%Y-%m-%d") for d in series[0][1].index]
we.cell(1,1,"Date")
for j,(nm,s) in enumerate(series,2): we.cell(1,j,nm)
for i,d in enumerate(ddt,2):
    we.cell(i,1,d)
    for j,(nm,s) in enumerate(series,2): we.cell(i,j,round(float(s.iloc[i-2]),2))
for cc in we[1]: cc.font=Font(bold=True)
last=len(ddt)+1
lc=LineChart(); lc.title="ORIGINAL 8B variants — equity @50bp (log)"; lc.y_axis.scaling.logBase=10; lc.height=12; lc.width=26
for j in range(2,2+len(series)): lc.add_data(Reference(we,min_col=j,min_row=1,max_row=last),titles_from_data=True)
lc.set_categories(Reference(we,min_col=1,min_row=2,max_row=last)); we.add_chart(lc, get_column_letter(len(series)+3)+"2")

wb.save(OUT); print("\nsaved", OUT, "| sheets:", wb.sheetnames)
