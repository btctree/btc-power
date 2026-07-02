"""Compare 1x (no lev) vs 2x vs 3x vs CONDITIONAL-leverage B model, with full trade logs + charts.

Conditional leverage = different leverage by DIRECTION (long vs short), REGIME/strategy, and
CONFIDENCE (|ensemble|) — principled (NOT tuned), grounded in earlier findings:
 * longs are more reliable than shorts (squeeze risk)  -> base long 3x, short 2x
 * trends are more reliable than chop                   -> regime factor 1.0 .. 0.5
 * higher conviction deserves more size                 -> confidence factor 1.0 / 0.85 / 0.7
All models: same v2 ensemble, conviction filter, vol-target 60%, DD-kill 30%, turnover control
(EMA-smooth 5 + 0.15 deadband). Honest: 2014+, fees, slippage on turnover, intraday liquidation.
"""
import os
import numpy as np, pandas as pd
import stable_combo as sc
import live_engine as le
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ANN = 365
OUT = r"C:\Users\user\OneDrive\Desktop\New setup for BTC\BTC_leverage_compare.xlsx"

df, reg0, memb = sc.prep()
reg, emap, exp_raw = le.ensemble_ctx(df, memb)
expf = np.where(np.abs(exp_raw) >= 0.4, exp_raw, 0.0)
close = df["close"].to_numpy(); low = df["low"].to_numpy(); high = df["high"].to_numpy()
dates = pd.to_datetime(df["Date"])
rv = pd.Series(close).pct_change().rolling(20).std().to_numpy() * np.sqrt(ANN)
n = len(df); i0 = 260
e_sm = pd.Series(expf).ewm(span=5, adjust=False).mean().to_numpy()   # turnover-smoothed exposure

# ---- conditional leverage rule ----
REGF = {"STRONG_UP": 1.0, "TREND_UP": 1.0, "PULLBACK_UP": 0.7, "BOUNCE_DOWN": 0.7,
        "STRONG_DOWN": 0.8, "TREND_DOWN": 0.9, "CHOP_HIVOL": 0.5, "RANGE": 0.5, "NEUTRAL": 0.5}
def conf_f(a):
    a = abs(a)
    return 1.0 if a >= 0.75 else (0.85 if a >= 0.5 else 0.7)
def cond_lev_arr():
    lv = np.zeros(n)
    for i in range(n):
        s = e_sm[i]
        if s == 0:
            continue
        base = 3.0 if s > 0 else 2.0
        lv[i] = base * REGF.get(reg[i], 0.5) * conf_f(s)
    return lv

def sim(lev_arr, slip=0.005, vol_target=0.60, dd_kill=0.30, band=0.15, fee=0.0005, maint=0.01):
    equity = peak = 500.0; held = 0.0; liqs = 0; turn_tot = 0.0
    eq = np.full(n, 500.0); heldarr = np.zeros(n)
    for i in range(i0, n):
        L = lev_arr[i - 1]
        tgt = e_sm[i - 1] * L
        if vol_target > 0 and rv[i - 1] == rv[i - 1] and rv[i - 1] > 0:
            tgt = np.clip(tgt, -abs(e_sm[i - 1]) * L, abs(e_sm[i - 1]) * L)
            tgt *= min(1.0, vol_target / rv[i - 1])
        if dd_kill > 0 and equity < peak * (1 - dd_kill):
            tgt *= 0.5
        e = tgt
        if band > 0 and abs(e - held) < band and not (e == 0 and held != 0):
            e = held
        adverse = (-(low[i] / close[i - 1] - 1)) if e > 0 else ((high[i] / close[i - 1] - 1) if e < 0 else 0.0)
        if e != 0 and abs(e) * max(adverse, 0) >= (1 - maint):
            equity *= 0.01; liqs += 1; held = 0.0; eq[i] = equity; heldarr[i] = 0.0; peak = max(peak, equity); continue
        ret = close[i] / close[i - 1] - 1
        turn = abs(e - held); turn_tot += turn
        equity *= (1 + e * ret); equity -= equity * turn * (fee + slip)
        held = e; eq[i] = equity; heldarr[i] = e; peak = max(peak, equity)
    return eq[i0:], heldarr[i0:], liqs, turn_tot / ((n - i0) / ANN)

def met(eq):
    eq = np.asarray(eq, float); r = np.diff(eq) / eq[:-1]; r = r[np.isfinite(r)]
    yrs = len(eq) / ANN
    cagr = (eq[-1] / 500) ** (1 / yrs) - 1 if eq[-1] > 0 else -1
    sh = r.mean() * ANN / (r.std(ddof=1) * np.sqrt(ANN)) if r.std() > 0 else float("nan")
    dn = r[r < 0].std(ddof=1) * np.sqrt(ANN) if (r < 0).sum() > 1 else float("nan")
    so = r.mean() * ANN / dn if dn and dn > 0 else float("nan")
    dd = (eq / np.maximum.accumulate(eq) - 1).min()
    return dict(final=eq[-1], pnl=eq[-1] - 500, cagr=cagr, sharpe=sh, sortino=so, maxdd=dd,
                calmar=cagr / abs(dd) if dd < 0 else float("nan"))

WIN = [("W1 2018", "2017-12-16", "2018-11-18"), ("W2 2021-22", "2021-10-20", "2022-03-09"),
       ("W3 2025", "2025-05-22", "2025-12-01")]
ddt_full = dates.iloc[i0:].reset_index(drop=True)
def wdd(eq):
    s = pd.Series(eq, index=ddt_full); out = []
    for nm, a, b in WIN:
        w = s.loc[a:b]; out.append((w / w.cummax() - 1).min() * 100 if len(w) else float("nan"))
    return out

def trades_from(eq, held):
    cl = close[i0:]; rgs = reg[i0:]; dts = ddt_full.dt.strftime("%Y-%m-%d").tolist()
    tr = []; m = len(held); i = 0
    while i < m:
        s = 1 if held[i] > 0 else (-1 if held[i] < 0 else 0)
        if s == 0:
            i += 1; continue
        j = i
        while j + 1 < m and (1 if held[j + 1] > 0 else (-1 if held[j + 1] < 0 else 0)) == s:
            j += 1
        eq_start = eq[i - 1] if i > 0 else 500.0
        entry = float(cl[i]); exit_ = float(cl[j])
        tr.append(dict(entry_dt=dts[i], exit_dt=dts[j], days=j - i + 1,
                       direction="LONG" if s > 0 else "SHORT", regime=rgs[i],
                       strategy=", ".join(emap.get(rgs[i]) or []) or "-",
                       entry_px=round(entry, 2), exit_px=round(exit_, 2),
                       btc_move=round((exit_ / entry - 1) * s, 4),
                       avg_pos=round(float(np.mean(np.abs(held[i:j + 1]))), 2),
                       trade_pnl=round(eq[j] / eq_start - 1, 4)))
        i = j + 1
    return tr

# ---- run all models at 0bp and 50bp ----
condarr = cond_lev_arr()
MODELS = [("1x (no leverage)", np.full(n, 1.0)), ("2x B", np.full(n, 2.0)),
          ("3x B", np.full(n, 3.0)), ("Conditional", condarr)]
res = {}
for slip in (0.0, 0.005):
    for name, la in MODELS:
        eq, held, liq, turn = sim(la, slip=slip)
        res[(name, slip)] = dict(eq=eq, held=held, liq=liq, turn=turn, m=met(eq), wdd=wdd(eq))

eff_cond = np.mean(condarr[condarr > 0])
print(f"Conditional avg effective leverage when in-market: {eff_cond:.2f}x")
for slip in (0.0, 0.005):
    print(f"\n-- slippage {int(slip*10000)}bp --")
    for name, _ in MODELS:
        d = res[(name, slip)]; m = d["m"]; w = d["wdd"]
        print(f"  {name:18s} ${m['final']:>14,.0f}  CAGR {m['cagr']*100:>4.0f}%  Sh {m['sharpe']:.2f}  "
              f"Cal {m['calmar']:.2f}  DD {m['maxdd']*100:.0f}%  liq {d['liq']:>2d} | W {w[0]:.0f}/{w[1]:.0f}/{w[2]:.0f}%")

# ======================= EXCEL =======================
HDR = Font(bold=True, color="FFFFFF"); HFILL = PatternFill("solid", fgColor="1F3864")
TITLE = Font(bold=True, size=14); SUB = Font(italic=True, color="666666")
thin = Side(style="thin", color="DDDDDD"); BORD = Border(thin, thin, thin, thin)
wb = Workbook()

# ---- Summary ----
ws = wb.active; ws.title = "Summary"
ws["A1"] = "BTC B-model — leverage comparison (none / 2x / 3x / conditional)"; ws["A1"].font = TITLE
ws["A2"] = ("All = same v2 ensemble + conviction filter + vol-target 60% + DD-kill 30% + turnover control. "
            "2014+, $500 start, 5bp fees, slippage on turnover, honest intraday liquidation.")
ws["A2"].font = SUB
ws["A4"] = "Conditional-leverage rule (principled, not tuned):"; ws["A4"].font = Font(bold=True)
rule = [f"  base by direction: LONG 3x, SHORT 2x (shorts riskier)",
        f"  x regime factor: STRONG/TREND_UP 1.0 | DOWN-trends 0.8-0.9 | pullback/bounce 0.7 | CHOP/RANGE 0.5",
        f"  x confidence factor (|ensemble|): >=0.75 ->1.0 | 0.5-0.75 ->0.85 | 0.4-0.5 ->0.7",
        f"  => realized avg effective leverage in-market = {eff_cond:.2f}x"]
for k, t in enumerate(rule):
    ws.cell(5 + k, 1, t)
r0 = 10
cols = ["Model", "Slippage", "Final $", "Total PnL $", "CAGR", "Sharpe", "Sortino", "Calmar",
        "max DD", "Liq", "Turn/yr", "W1 2018", "W2 2021-22", "W3 2025"]
for j, c in enumerate(cols, 1):
    cell = ws.cell(r0, j, c); cell.font = HDR; cell.fill = HFILL; cell.alignment = Alignment(horizontal="center")
rr = r0
for slip in (0.0, 0.005):
    for name, _ in MODELS:
        rr += 1; d = res[(name, slip)]; m = d["m"]; w = d["wdd"]
        vals = [name, f"{int(slip*10000)}bp", round(m["final"]), round(m["pnl"]), m["cagr"], m["sharpe"],
                m["sortino"], m["calmar"], m["maxdd"], d["liq"], round(d["turn"], 1),
                w[0] / 100, w[1] / 100, w[2] / 100]
        for j, v in enumerate(vals, 1):
            c = ws.cell(rr, j, v); c.border = BORD
        ws.cell(rr, 3).number_format = '#,##0'; ws.cell(rr, 4).number_format = '#,##0'
        for jj in (5, 9, 12, 13, 14):
            ws.cell(rr, jj).number_format = '0%'
        for jj in (6, 7, 8):
            ws.cell(rr, jj).number_format = '0.00'
        if name == "Conditional":
            for j in range(1, 15):
                ws.cell(rr, j).fill = PatternFill("solid", fgColor="FFF2CC")
ws.column_dimensions["A"].width = 18
for col in "BCDEFGHIJKLMN":
    ws.column_dimensions[col].width = 12

# ---- Equity compare ----
we = wb.create_sheet("Equity Compare")
ddt = ddt_full.dt.strftime("%Y-%m-%d").tolist()
hdr = ["Date", "BTC close"] + [f"{nm} @0bp" for nm, _ in MODELS] + [f"{nm} @50bp" for nm, _ in MODELS]
for j, h in enumerate(hdr, 1):
    c = we.cell(1, j, h); c.font = Font(bold=True)
for i, dstr in enumerate(ddt, 2):
    we.cell(i, 1, dstr); we.cell(i, 2, round(float(close[i0:][i - 2]), 2))
    for k, (nm, _) in enumerate(MODELS):
        we.cell(i, 3 + k, round(float(res[(nm, 0.0)]["eq"][i - 2]), 2))
        we.cell(i, 7 + k, round(float(res[(nm, 0.005)]["eq"][i - 2]), 2))
last = len(ddt) + 1
cats = Reference(we, min_col=1, min_row=2, max_row=last)
def chart(anchor, title, col_start):
    ch = LineChart(); ch.title = title; ch.style = 2
    ch.y_axis.title = "Equity $ (log)"; ch.x_axis.title = "Date"; ch.height = 11; ch.width = 26
    ch.y_axis.scaling.logBase = 10
    for k in range(4):
        ref = Reference(we, min_col=col_start + k, min_row=1, max_row=last)
        ch.add_data(ref, titles_from_data=True)
    ch.set_categories(cats); we.add_chart(ch, anchor)
endcol = get_column_letter(len(hdr) + 2)
chart(f"{endcol}2", "Equity @ 50bp (realistic): 1x / 2x / 3x / Conditional", 7)
chart(f"{endcol}24", "Equity @ 0bp (best case): 1x / 2x / 3x / Conditional", 3)

# ---- trade sheets ----
def trade_sheet(title, name):
    wsx = wb.create_sheet(title)
    tr = trades_from(res[(name, 0.005)]["eq"], res[(name, 0.005)]["held"])
    hd = ["Entry", "Exit", "Days", "Direction", "Market type", "Strategy (engines)",
          "Entry $", "Exit $", "BTC move", "Avg pos (x)", "Trade P&L (lev)"]
    for j, h in enumerate(hd, 1):
        c = wsx.cell(1, j, h); c.font = HDR; c.fill = HFILL
    for i, t in enumerate(tr, 2):
        row = [t["entry_dt"], t["exit_dt"], t["days"], t["direction"], t["regime"], t["strategy"],
               t["entry_px"], t["exit_px"], t["btc_move"], t["avg_pos"], t["trade_pnl"]]
        for j, v in enumerate(row, 1):
            wsx.cell(i, j, v)
        wsx.cell(i, 9).number_format = '0.0%'; wsx.cell(i, 11).number_format = '0.0%'
    wsx.column_dimensions["A"].width = 12; wsx.column_dimensions["B"].width = 12
    wsx.column_dimensions["F"].width = 20
    return len(tr)
n2 = trade_sheet("Trades 2x", "2x B")
n3 = trade_sheet("Trades 3x", "3x B")
nc = trade_sheet("Trades Conditional", "Conditional")

wb.save(OUT)
print(f"\nsaved {OUT}")
print(f"trades: 2x={n2}  3x={n3}  conditional={nc}")
print("sheets:", wb.sheetnames)
