"""Generate a detailed Excel trade report for the BTC Power signal.
Runs the production engine (regime-switch, let-winners-run, 10%/7% trailing, asymmetric
shorts) on Binance daily decisions with 1-MINUTE intraday fills (real minute timestamps
on stop exits). Logs every trade with market type / strategy / confidence / cut-loss /
sizing / margin, and PnL both WITHOUT margin (spot 1x) and WITH margin (long 5x / short 2x).

Outputs: New setup for BTC/BTC_Power_trade_report.xlsx
Sheets: Trades, Summary, By Year, Daily.
"""
import os, datetime as dt
import numpy as np
import pandas as pd
import regime_switch_sim as rss

HERE = os.path.dirname(__file__)
DATA = os.path.join(HERE, "..", "data")
OUTXLSX = os.path.join(HERE, "..", "..", "BTC_Power_trade_report.xlsx")

MAP = rss.MAP
TREND_GROUP = rss.TREND_GROUP
BEAR_GROUP = rss.BEAR_GROUP
SHORT_REGIMES = {"CHOP_HIGHVOL", "BEAR_TREND"}
CUT_L, CUT_S = 0.10, 0.07
LEV_L, LEV_S = 5.0, 2.0            # "with margin" leverage (long / short)
FEE = 0.0005
EPOCH = dt.datetime(1970, 1, 1)


def conf_bucket(cs):
    return ("High", 1.0) if cs >= 1.8 else (("Med", 0.7) if cs >= 1.0 else ("Low", 0.4))


def suited(pos_kind, pos_regime, rg):
    if pos_kind == "trend" and pos_regime in TREND_GROUP:
        return rg in TREND_GROUP
    if pos_regime in BEAR_GROUP:
        return rg in BEAR_GROUP
    return rg == pos_regime


def ts(day_ord, minute):
    return (EPOCH + dt.timedelta(days=int(day_ord), minutes=int(minute))).strftime("%Y-%m-%d %H:%M")


def run():
    df, reg, memb, intr, fund = rss.load()
    close = df["close"].to_numpy(); dates = df["Date"].tolist(); n = len(df)
    # 1-min with minute timestamps
    z = np.load(os.path.join(DATA, "intraday_1m.npz"))
    day = z["day"]; minute = z["minute"]; H = z["h"]; L = z["l"]; C = z["c"]
    ud, st = np.unique(day, return_index=True); en = np.append(st[1:], len(day))
    di = {int(d): (int(s), int(e)) for d, s, e in zip(ud, st, en)}

    def dord(s):
        y, m, d = map(int, s.split("-")); return (dt.date(y, m, d) - dt.date(1970, 1, 1)).days

    eq_nm = eq_wm = 500.0
    pos = None
    trades = []
    daily = []
    start_i = 260

    for i in range(n):
        date_i = dates[i]; rg = reg[i]
        o = dord(date_i); sl = di.get(o)
        # ---- manage open position intraday ----
        if pos is not None:
            d = pos["dir"]; cl = pos["cl"]; exited = False; fillp = None; exmin = 0; reason = None
            if sl:
                s0, e0 = sl
                lows = L[s0:e0]; highs = H[s0:e0]; mins = minute[s0:e0]
                for k in range(len(lows)):
                    bl = lows[k]; bh = highs[k]
                    if d == 1:
                        if bh > pos["hi"]:
                            pos["hi"] = bh; pos["stop"] = max(pos["stop"], pos["hi"] * (1 - cl))
                        if bl <= pos["stop"]:
                            fillp = pos["stop"]; exmin = mins[k]; reason = "Trail/Cut-loss"; exited = True; break
                    else:
                        if bl < pos["lo"]:
                            pos["lo"] = bl; pos["stop"] = min(pos["stop"], pos["lo"] * (1 + cl))
                        if bh >= pos["stop"]:
                            fillp = pos["stop"]; exmin = mins[k]; reason = "Trail/Cut-loss"; exited = True; break
            if not exited:
                # signal / regime exit at daily close
                m = memb[pos["strat"]][i]
                sigx = (m * pos["dir"] < 0) if pos["kind"] == "trend" else (m * pos["dir"] <= 0)
                su = suited(pos["kind"], pos["regime"], rg)
                if (not su) or sigx:
                    fillp = close[i]; exmin = 0; reason = "Regime-change" if not su else "Signal-exit"
                    exited = True
            if exited:
                ret = (fillp / pos["entry"] - 1.0) * pos["dir"]
                # no-margin
                fee_nm = pos["notional_nm"] * FEE * 2
                pnl_nm = pos["notional_nm"] * ret - fee_nm
                eq_nm += pnl_nm
                # with-margin
                fee_wm = pos["notional_wm"] * FEE * 2
                pnl_wm = pos["notional_wm"] * ret - fee_wm
                eq_wm += pnl_wm
                trades.append(dict(
                    entry_dt=pos["entry_dt"], exit_dt=ts(o, exmin) if reason == "Trail/Cut-loss" else date_i + " 00:00",
                    days=i - pos["entry_i"], market=pos["regime"], strategy=pos["strat"],
                    direction="LONG" if pos["dir"] == 1 else "SHORT",
                    confidence=pos["bucket"], conf_score=round(pos["conf"], 2),
                    entry=round(pos["entry"], 2), exit=round(fillp, 2), ret=ret,
                    cutloss_lvl=round(pos["cutloss_lvl"], 2), cutloss_pct=(CUT_L if pos["dir"] == 1 else CUT_S),
                    sizing=pos["deploy"], margin=round(pos["margin_wm"], 2), lev=pos["lev"],
                    pnl_nm=round(pnl_nm, 2), pnl_wm=round(pnl_wm, 2), reason=reason,
                    eq_nm=round(eq_nm, 2), eq_wm=round(eq_wm, 2)))
                pos = None
        # ---- entry at close (flat only) ----
        if pos is None and i >= start_i and eq_nm > 1:
            strat, kind, cs = MAP.get(rg, (None, "flat", 0))
            if strat and cs >= 0.5:
                m = memb[strat][i]
                d = 1 if m > 0 else (-1 if m < 0 else 0)
                if d == -1 and rg not in SHORT_REGIMES:
                    d = 0
                if d != 0:
                    bk, dep = conf_bucket(cs)
                    if d == -1:
                        dep *= 0.5
                    cl = CUT_L if d == 1 else CUT_S
                    lev = LEV_L if d == 1 else LEV_S
                    entry = close[i]
                    notional_nm = dep * eq_nm            # 1x spot
                    margin_wm = dep * eq_wm
                    notional_wm = margin_wm * lev         # leveraged
                    eq_nm -= notional_nm * FEE            # entry fee already counted in *2 at close; keep simple: don't double. Remove:
                    eq_nm += notional_nm * FEE            # (no-op to avoid double fee; fees charged at close as *2)
                    cutloss_lvl = entry * (1 - cl) if d == 1 else entry * (1 + cl)
                    stop = cutloss_lvl
                    pos = dict(entry=entry, entry_i=i, entry_dt=date_i + " 00:00", dir=d, strat=strat,
                               kind=kind, regime=rg, bucket=bk, conf=cs, deploy=dep, cl=cl, lev=lev,
                               hi=entry, lo=entry, stop=stop, cutloss_lvl=cutloss_lvl,
                               notional_nm=notional_nm, notional_wm=notional_wm, margin_wm=margin_wm)
        # daily snapshot
        held = "FLAT" if pos is None else ("LONG" if pos["dir"] == 1 else "SHORT")
        strat_d, kind_d, cs_d = MAP.get(rg, (None, "flat", 0))
        daily.append(dict(date=date_i, close=round(close[i], 2), market=rg,
                          active_strategy=(strat_d if (strat_d and cs_d >= 0.5) else "STAND ASIDE"),
                          position=held, conf_score=round(cs_d, 2) if strat_d else 0,
                          rsi=round(float(df["RSI"].iloc[i]), 1) if df["RSI"].iloc[i] == df["RSI"].iloc[i] else None))
    return trades, daily


def to_excel(trades, daily):
    tdf = pd.DataFrame(trades)
    tdf.insert(0, "#", range(1, len(tdf) + 1))
    cols = {"#": "#", "entry_dt": "Entry Date/Time", "exit_dt": "Exit Date/Time", "days": "Days Held",
            "market": "Market Type", "strategy": "Strategy", "direction": "Direction",
            "confidence": "Confidence", "conf_score": "Conf Score", "entry": "Entry $", "exit": "Exit $",
            "ret": "Return %", "cutloss_lvl": "Cut-Loss $", "cutloss_pct": "Cut-Loss %",
            "sizing": "Sizing % equity", "margin": "Margin $ used", "lev": "Leverage (margin)",
            "pnl_nm": "PnL no-margin $", "pnl_wm": "PnL with-margin $", "reason": "Exit Reason",
            "eq_nm": "Equity no-margin $", "eq_wm": "Equity margin $"}
    tdf = tdf[list(cols)].rename(columns=cols)

    # Summary
    wins = (tdf["Return %"] > 0).sum(); ntr = len(tdf); losses = ntr - wins
    longw = ((tdf["Direction"] == "LONG") & (tdf["Return %"] > 0)).sum(); longn = (tdf["Direction"] == "LONG").sum()
    shortw = ((tdf["Direction"] == "SHORT") & (tdf["Return %"] > 0)).sum(); shortn = (tdf["Direction"] == "SHORT").sum()
    gw = tdf.loc[tdf["Return %"] > 0, "Return %"].sum(); gl = -tdf.loc[tdf["Return %"] < 0, "Return %"].sum()
    summary = pd.DataFrame({
        "Metric": ["Total trades", "Wins", "Losses", "Win ratio (overall)", "Win ratio LONG", "Win ratio SHORT",
                   "Avg win %", "Avg loss %", "Profit factor", "Final equity NO-margin $",
                   "Final equity WITH-margin $", "Start $", "Period"],
        "Value": [ntr, wins, losses, f"{wins/ntr*100:.1f}%",
                  f"{(longw/longn*100 if longn else 0):.1f}% ({longw}/{longn})",
                  f"{(shortw/shortn*100 if shortn else 0):.1f}% ({shortw}/{shortn})",
                  f"{tdf.loc[tdf['Return %']>0,'Return %'].mean()*100:.2f}%",
                  f"{tdf.loc[tdf['Return %']<=0,'Return %'].mean()*100:.2f}%",
                  f"{(gw/gl if gl else 0):.2f}",
                  f"{tdf['Equity no-margin $'].iloc[-1]:,.0f}",
                  f"{tdf['Equity margin $'].iloc[-1]:,.0f}", "500",
                  f"{tdf['Entry Date/Time'].iloc[0][:10]} .. {tdf['Exit Date/Time'].iloc[-1][:10]}"]})
    notes = pd.DataFrame({"Metric": [
        "", "NOTES",
        "PnL no-margin = SPOT 1x = realistic, tradeable result.",
        "PnL with-margin = LONG 5x / SHORT 2x, sizing as margin. OPTIMISTIC:",
        "  assumes the 7-10% trailing stop fills perfectly intraday; under real",
        "  crash slippage the leveraged result collapses (see report Truth tab).",
        "Sizing: confidence-scaled (High 100% / Med 70% of equity); shorts half-size.",
        "Cut-loss: 10% trailing (long) / 7% (short), ratchets behind high/low-water.",
        "Backtest starts ~2018-05 after 200-day + 1-year indicator warmup.",
        "Hypothetical backtest, Binance BTCUSDT, 1-min intraday fills, 5bp/side fees.",
        "Not financial advice."],
        "Value": [""] * 11})
    summary = pd.concat([summary, notes], ignore_index=True)

    # By year (by exit year)
    tdf["_yr"] = tdf["Exit Date/Time"].str[:4]
    yr_rows = []
    for y, g in tdf.groupby("_yr"):
        w = (g["Return %"] > 0).sum()
        yr_rows.append(dict(Year=y, Trades=len(g), Wins=w, **{"Win %": f"{w/len(g)*100:.1f}%"},
                            **{"PnL no-margin $": round(g["PnL no-margin $"].sum(), 0),
                               "PnL with-margin $": round(g["PnL with-margin $"].sum(), 0),
                               "End equity (no-margin) $": round(g["Equity no-margin $"].iloc[-1], 0)}))
    ydf = pd.DataFrame(yr_rows)
    tdf = tdf.drop(columns=["_yr"])
    ddf = pd.DataFrame(daily).rename(columns={"date": "Date", "close": "Close $", "market": "Market Type",
                                              "active_strategy": "Active Strategy", "position": "Position",
                                              "conf_score": "Conf Score", "rsi": "RSI"})

    with pd.ExcelWriter(OUTXLSX, engine="openpyxl") as xw:
        summary.to_excel(xw, sheet_name="Summary", index=False)
        ydf.to_excel(xw, sheet_name="By Year", index=False)
        tdf.to_excel(xw, sheet_name="Trades", index=False)
        ddf.to_excel(xw, sheet_name="Daily", index=False)
        _format(xw, tdf, summary, ydf, ddf)
    return tdf, summary, ydf


def _format(xw, tdf, summary, ydf, ddf):
    from openpyxl.styles import Font, PatternFill, Alignment
    hdr_fill = PatternFill("solid", fgColor="1F2A44"); hdr_font = Font(bold=True, color="FFFFFF")
    for name in xw.book.sheetnames:
        ws = xw.book[name]
        ws.freeze_panes = "A2"
        for c in ws[1]:
            c.fill = hdr_fill; c.font = hdr_font; c.alignment = Alignment(horizontal="center")
        ws.auto_filter.ref = ws.dimensions
        for col in ws.columns:
            w = max((len(str(c.value)) if c.value is not None else 0) for c in col)
            ws.column_dimensions[col[0].column_letter].width = min(max(w + 2, 10), 22)
    # number formats on Trades
    ws = xw.book["Trades"]
    hdrs = [c.value for c in ws[1]]
    fmt = {"Return %": "0.00%", "Cut-Loss %": "0.0%", "Sizing % equity": "0%",
           "Entry $": "#,##0.00", "Exit $": "#,##0.00", "Cut-Loss $": "#,##0.00",
           "Margin $ used": "#,##0", "PnL no-margin $": "#,##0;[Red]-#,##0",
           "PnL with-margin $": "#,##0;[Red]-#,##0", "Equity no-margin $": "#,##0", "Equity margin $": "#,##0",
           "Leverage (margin)": '0"x"'}
    for j, h in enumerate(hdrs, 1):
        if h in fmt:
            for r in range(2, ws.max_row + 1):
                ws.cell(row=r, column=j).number_format = fmt[h]


if __name__ == "__main__":
    trades, daily = run()
    tdf, summary, ydf = to_excel(trades, daily)
    print(f"Trades: {len(tdf)}  -> {OUTXLSX}")
    print(summary.to_string(index=False))
    print("\nBy year:")
    print(ydf.to_string(index=False))
