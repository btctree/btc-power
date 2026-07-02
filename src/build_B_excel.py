"""Full Excel for the 'B: turnover-control' model (the free-lunch fix) + leveraged versions
at 0 / 50 / 100 bp slippage, with charts.

B model = conviction-filtered v2 ensemble, vol_target 0.60, DD-kill 0.30, PLUS turnover control
(EMA smooth span=5 + 0.15 deadband). Honest: real data 2014+, 5bp fees, slippage on turnover,
honest intraday liquidation at leverage. No funding (spot/cash-margin basis) to reconcile with the
$63,546 headline. 1x@30bp must reproduce ~$63,546.
"""
import os
import numpy as np, pandas as pd
import stable_combo as sc
import live_engine as le
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ANN = 365
OUTPATH = r"C:\Users\user\OneDrive\Desktop\New setup for BTC\BTC_B_turnover_model.xlsx"

df, reg0, memb = sc.prep()
reg, emap, exp_raw = le.ensemble_ctx(df, memb)
expf = np.where(np.abs(exp_raw) >= 0.4, exp_raw, 0.0)
close = df["close"].to_numpy(); low = df["low"].to_numpy(); high = df["high"].to_numpy()
dates = pd.to_datetime(df["Date"])
rv = pd.Series(close).pct_change().rolling(20).std().to_numpy() * np.sqrt(ANN)
n = len(df); i0 = 260


def isim(exp, lev=1.0, slip=0.0, smooth=5, band=0.15, vol_target=0.60, dd_kill=0.30,
         fee=0.0005, maint=0.01):
    """Instrumented mirror of stable_combo.simulate (no funding) -> eq, liqs, per-day arrays."""
    e_in = pd.Series(exp).ewm(span=smooth, adjust=False).mean().to_numpy() if smooth > 1 else exp.copy()
    equity = peak = 500.0; held = 0.0; liqs = 0; turn_tot = 0.0
    eq = np.full(n, 500.0); used = np.zeros(n); rets = np.zeros(n)
    for i in range(i0, n):
        tgt = e_in[i - 1] * lev
        if vol_target > 0 and rv[i - 1] == rv[i - 1] and rv[i - 1] > 0:
            tgt = np.clip(tgt, -abs(e_in[i - 1]) * lev, abs(e_in[i - 1]) * lev)
            tgt *= min(1.0, vol_target / rv[i - 1])
        if dd_kill > 0 and equity < peak * (1 - dd_kill):
            tgt *= 0.5
        e = tgt
        if band > 0 and abs(e - held) < band and not (e == 0 and held != 0):
            e = held
        if e > 0:
            adverse = -(low[i] / close[i - 1] - 1)
        elif e < 0:
            adverse = (high[i] / close[i - 1] - 1)
        else:
            adverse = 0.0
        if e != 0 and abs(e) * max(adverse, 0) >= (1 - maint):
            equity *= 0.01; liqs += 1; held = 0.0
            eq[i] = equity; used[i] = 0.0; peak = max(peak, equity); continue
        ret = close[i] / close[i - 1] - 1
        turn = abs(e - held); turn_tot += turn
        equity *= (1 + e * ret); equity -= equity * turn * (fee + slip)
        held = e; eq[i] = equity; peak = max(peak, equity)
        used[i] = e; rets[i] = ret
    return eq, liqs, used, rets, turn_tot / ((n - i0) / ANN)


def metrics(eq):
    eq = np.asarray(eq, float); r = np.diff(eq) / eq[:-1]; r = r[np.isfinite(r)]
    yrs = len(eq) / ANN
    cagr = (eq[-1] / 500) ** (1 / yrs) - 1 if eq[-1] > 0 else -1
    sh = r.mean() * ANN / (r.std(ddof=1) * np.sqrt(ANN)) if r.std() > 0 else float("nan")
    dn = r[r < 0].std(ddof=1) * np.sqrt(ANN) if (r < 0).sum() > 1 else float("nan")
    so = r.mean() * ANN / dn if dn and dn > 0 else float("nan")
    dd = (eq / np.maximum.accumulate(eq) - 1).min()
    return dict(final=eq[-1], pnl=eq[-1] - 500, cagr=cagr, sharpe=sh, sortino=so,
                maxdd=dd, calmar=cagr / abs(dd) if dd < 0 else float("nan"))


SCEN = [
    ("1x · 30bp (headline B)", 1, 0.003),
    ("1x · 0bp", 1, 0.0),
    ("2x · 0bp", 2, 0.0), ("2x · 50bp", 2, 0.005), ("2x · 100bp", 2, 0.010),
    ("3x · 0bp", 3, 0.0), ("3x · 50bp", 3, 0.005), ("3x · 100bp", 3, 0.010),
    ("5x · 0bp", 5, 0.0), ("5x · 50bp", 5, 0.005), ("5x · 100bp", 5, 0.010),
]
res = {}
for name, lev, slip in SCEN:
    eq, liq, used, rets, turn = isim(expf, lev=lev, slip=slip)
    res[name] = dict(eq=eq[i0:], liq=liq, used=used[i0:], rets=rets[i0:], turn=turn,
                     lev=lev, slip=slip, m=metrics(eq[i0:]))
print(f"RECONCILE B 1x@30bp final = ${res['1x · 30bp (headline B)']['m']['final']:,.0f} (target ~$63,546)")
for name in [s[0] for s in SCEN]:
    m = res[name]["m"]
    print(f"  {name:24s} ${m['final']:>14,.0f}  CAGR {m['cagr']*100:>5.0f}%  Sh {m['sharpe']:.2f}  "
          f"Cal {m['calmar']:.2f}  DD {m['maxdd']*100:.0f}%  liq {res[name]['liq']}")

# ---------------- build workbook ----------------
ddt = dates.iloc[i0:].dt.strftime("%Y-%m-%d").tolist()
btc = close[i0:]
wb = Workbook()
HDR = Font(bold=True, color="FFFFFF"); HFILL = PatternFill("solid", fgColor="1F3864")
TITLE = Font(bold=True, size=14); SUB = Font(italic=True, color="666666")
thin = Side(style="thin", color="DDDDDD"); BORD = Border(*([thin] * 4))

# ---- Summary ----
ws = wb.active; ws.title = "Summary"
ws["A1"] = "BTC 'B' model — turnover-controlled ensemble + leverage/slippage"; ws["A1"].font = TITLE
ws["A2"] = ("B = conviction-filtered regime ensemble · vol-target 60% · DD-kill 30% · EMA-smooth(5) + 0.15 deadband. "
            "2014+, $500 start, 5bp fees, slippage on turnover, honest intraday liquidation. No funding.")
ws["A2"].font = SUB
cols = ["Scenario", "Leverage", "Slippage", "Final $", "Total PnL $", "CAGR", "Sharpe", "Sortino",
        "Calmar", "max DD", "Liquidations", "Turnover/yr"]
r0 = 4
for j, c in enumerate(cols, 1):
    cell = ws.cell(r0, j, c); cell.font = HDR; cell.fill = HFILL; cell.alignment = Alignment(horizontal="center")
for k, name in enumerate([s[0] for s in SCEN], 1):
    d = res[name]; m = d["m"]; rr = r0 + k
    vals = [name, f"{d['lev']}x", f"{int(d['slip']*10000)}bp", round(m["final"]), round(m["pnl"]),
            m["cagr"], m["sharpe"], m["sortino"], m["calmar"], m["maxdd"], d["liq"], round(d["turn"], 1)]
    for j, v in enumerate(vals, 1):
        cell = ws.cell(rr, j, v); cell.border = BORD
    ws.cell(rr, 4).number_format = '#,##0'; ws.cell(rr, 5).number_format = '#,##0'
    ws.cell(rr, 6).number_format = '0%'; ws.cell(rr, 10).number_format = '0%'
    for jj in (7, 8, 9):
        ws.cell(rr, jj).number_format = '0.00'
    if k == 1:
        for j in range(1, 13):
            ws.cell(rr, j).fill = PatternFill("solid", fgColor="E2EFDA")
ws.column_dimensions["A"].width = 24
for col in "BCDEFGHIJKL":
    ws.column_dimensions[col].width = 13

# ---- Equity Curves ----
we = wb.create_sheet("Equity Curves")
we.cell(1, 1, "Date"); we.cell(1, 2, "BTC close")
names = [s[0] for s in SCEN]
for j, name in enumerate(names, 3):
    we.cell(1, j, name)
for i, dstr in enumerate(ddt, 2):
    we.cell(i, 1, dstr); we.cell(i, 2, round(float(btc[i - 2]), 2))
    for j, name in enumerate(names, 3):
        we.cell(i, j, round(float(res[name]["eq"][i - 2]), 2))
for c in we.iter_rows(min_row=1, max_row=1):
    for cell in c:
        cell.font = Font(bold=True)
last = len(ddt) + 1
cats = Reference(we, min_col=1, min_row=2, max_row=last)


def addchart(anchor, title, colnames, logscale=True):
    ch = LineChart(); ch.title = title; ch.style = 2
    ch.y_axis.title = "Equity $ (from $500)"; ch.x_axis.title = "Date"
    ch.height = 11; ch.width = 26
    if logscale:
        ch.y_axis.scaling.logBase = 10
    ch.x_axis.number_format = "yyyy"; ch.x_axis.majorTimeUnit = "years"
    for name in colnames:
        col = 3 + names.index(name)
        ref = Reference(we, min_col=col, min_row=1, max_row=last)
        ch.add_data(ref, titles_from_data=True)
    ch.set_categories(cats)
    we.add_chart(ch, anchor)


col_letter = get_column_letter(len(names) + 4)
addchart(f"{col_letter}2", "Leverage @ 0bp slippage (best case)",
         ["1x · 0bp", "2x · 0bp", "3x · 0bp", "5x · 0bp"])
addchart(f"{col_letter}24", "5x — slippage sensitivity (0 / 50 / 100bp)",
         ["5x · 0bp", "5x · 50bp", "5x · 100bp"])
addchart(f"{col_letter}46", "2x — slippage sensitivity (0 / 50 / 100bp)",
         ["2x · 0bp", "2x · 50bp", "2x · 100bp"])
addchart(f"{col_letter}68", "Headline B 1x (30bp) vs leverage @ 50bp",
         ["1x · 30bp (headline B)", "2x · 50bp", "3x · 50bp", "5x · 50bp"])

# ---- B 1x daily detail ----
wd = wb.create_sheet("B_1x_daily")
hd = ["Date", "BTC close", "Market type", "Raw exposure", "Position held (x)", "Daily ret", "Equity $", "Drawdown"]
for j, h in enumerate(hd, 1):
    cell = wd.cell(1, j, h); cell.font = HDR; cell.fill = HFILL
base = res["1x · 30bp (headline B)"]
eqb = base["eq"]; usedb = base["used"]; retsb = base["rets"]
peakv = np.maximum.accumulate(eqb)
regd = reg[i0:]
expfd = expf[i0:]
for i in range(len(ddt)):
    rr = i + 2
    wd.cell(rr, 1, ddt[i]); wd.cell(rr, 2, round(float(btc[i]), 2))
    wd.cell(rr, 3, regd[i]); wd.cell(rr, 4, round(float(expfd[i]), 3))
    wd.cell(rr, 5, round(float(usedb[i]), 3)); wd.cell(rr, 6, round(float(retsb[i]), 4))
    wd.cell(rr, 7, round(float(eqb[i]), 2)); wd.cell(rr, 8, round(float(eqb[i] / peakv[i] - 1), 4))
    wd.cell(rr, 6).number_format = '0.00%'; wd.cell(rr, 8).number_format = '0.0%'
    wd.cell(rr, 7).number_format = '#,##0'
wd.column_dimensions["A"].width = 12; wd.column_dimensions["C"].width = 13
ch2 = LineChart(); ch2.title = "B model 1x — equity ($500 start, 30bp)"; ch2.height = 10; ch2.width = 24
ch2.y_axis.scaling.logBase = 10; ch2.y_axis.title = "Equity $"
ref = Reference(wd, min_col=7, min_row=1, max_row=len(ddt) + 1)
ch2.add_data(ref, titles_from_data=True)
ch2.set_categories(Reference(wd, min_col=1, min_row=2, max_row=len(ddt) + 1))
wd.add_chart(ch2, "J2")

wb.save(OUTPATH)
print("\nsaved:", OUTPATH)
print("sheets:", wb.sheetnames)
