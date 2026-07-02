"""FULL comparison of EVERY model discussed, optimized with the size-control logic, + Excel + charts.
Columns: $@0bp / $@50bp / $@1x / CAGR / Calmar / Sharpe / maxDD / Win% / W1 / W2 / W3.
Plus a 'Rejected approaches' sheet (tested, no edge). Honest: 2014+, VOL+FUND gates, dd-kill, liquidation.
"""
import os
import numpy as np, pandas as pd
import stable_combo as sc
import live_engine as le
from openpyxl import Workbook
from openpyxl.chart import LineChart, BarChart, Reference
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
ANN = 365
OUT = r"C:\Users\user\OneDrive\Desktop\New setup for BTC\BTC_ALL_MODELS_compare.xlsx"
df, reg0, memb = sc.prep()
reg, emap, exp_raw = le.ensemble_ctx(df, memb)
expf = np.where(np.abs(exp_raw) >= 0.4, exp_raw, 0.0)
close = df["close"].to_numpy(); low = df["low"].to_numpy(); high = df["high"].to_numpy()
n = len(df); i0 = 260; dstr = df["Date"].tolist(); dates = pd.to_datetime(df["Date"])
rv = pd.Series(close).pct_change().rolling(20).std().to_numpy() * np.sqrt(ANN)
HERE = os.path.dirname(os.path.abspath(__file__))
def lmap(p, col):
    if not os.path.exists(p): return {}
    f = pd.read_csv(p); return dict(zip(f.iloc[:, 0].astype(str), f[col]))
funding = np.array([lmap(os.path.join(HERE, "..", "data", "funding.csv"), "funding_rate").get(d, np.nan) for d in dstr])
def trail_rank(a, w=365):
    return pd.Series(a).rolling(w, min_periods=60).apply(lambda x: (x.iloc[-1] >= x).mean(), raw=False).to_numpy()
vol_rank = trail_rank(rv); fund_rank = trail_rank(funding)
gl = np.ones(n); gs = np.ones(n)
for i in range(n):
    if vol_rank[i] == vol_rank[i] and vol_rank[i] > 0.85: gl[i] *= 0.5; gs[i] *= 0.5
    if fund_rank[i] == fund_rank[i]:
        if fund_rank[i] > 0.90: gl[i] *= 0.5
        if fund_rank[i] < 0.10: gs[i] *= 0.5
halv = [pd.Timestamp(x) for x in ["2012-11-28","2016-07-09","2020-05-11","2024-04-20"]]
dsh = np.array([min([(d-h).days for h in halv if h<=d] or [9999]) for d in dates],float)
phase = (dsh/1458.0) % 1.0; cyc = np.where((phase<0.42)|(phase>0.62),1.0,-1.0)
REGF={"STRONG_UP":1.0,"TREND_UP":1.0,"PULLBACK_UP":0.7,"BOUNCE_DOWN":0.7,"STRONG_DOWN":0.8,"TREND_DOWN":0.9,"CHOP_HIVOL":0.5,"RANGE":0.5,"NEUTRAL":0.5}
def conf_f(a): a=abs(a); return 1.0 if a>=0.75 else (0.85 if a>=0.5 else 0.7)
esm=pd.Series(expf).ewm(span=5,adjust=False).mean().to_numpy()
condcap=np.array([(3.0 if esm[i]>0 else 2.0)*REGF.get(reg[i],0.5)*conf_f(esm[i]) for i in range(n)])

def sim(signal, cap, vt, smooth, band, slip, gates=True, dd_kill=0.30, legacy=None, fee=0.0005, maint=0.01):
    e_in = pd.Series(signal).ewm(span=smooth, adjust=False).mean().to_numpy() if smooth>1 else np.array(signal,float)
    caparr = cap if hasattr(cap,'__len__') else None
    eqv=peak=500.0; held=0.0; eq=np.full(n,500.0); E=np.zeros(n); liq=0
    for i in range(i0,n):
        sig=e_in[i-1]; g=(gl[i-1] if sig>0 else (gs[i-1] if sig<0 else 1.0)) if gates else 1.0
        if not (rv[i-1]==rv[i-1] and rv[i-1]>0): eq[i]=eqv; E[i]=held; continue
        if legacy: mult=legacy*min(1.0,0.60/rv[i-1])
        else: mult=min(caparr[i-1] if caparr is not None else cap, vt/rv[i-1])
        e=sig*g*mult
        if dd_kill>0 and eqv<peak*(1-dd_kill): e*=0.5
        if band>0 and abs(e-held)<band and not(e==0 and held!=0): e=held
        adv=(-(low[i]/close[i-1]-1)) if e>0 else ((high[i]/close[i-1]-1) if e<0 else 0.0)
        if e!=0 and abs(e)*max(adv,0)>=(1-maint): eqv*=0.01; liq+=1; held=0.0; eq[i]=eqv; E[i]=0.0; peak=max(peak,eqv); continue
        eqv*=(1+e*(close[i]/close[i-1]-1)); eqv-=eqv*abs(e-held)*(fee+slip)
        held=e; eq[i]=max(eqv,1e-9); E[i]=e; peak=max(peak,eqv)
    return eq, E, liq
WIN=[("2017-12-16","2018-11-18"),("2021-10-20","2022-03-09"),("2025-05-22","2025-12-01")]
def metr(eq,E):
    s=pd.Series(eq[i0:],index=dates.iloc[i0:].values); r=s.pct_change().dropna(); yrs=len(s)/ANN
    cagr=(s.iloc[-1]/500)**(1/yrs)-1 if s.iloc[-1]>0 else -1
    sh=r.mean()*ANN/(r.std(ddof=1)*np.sqrt(ANN)) if r.std()>0 else float('nan')
    dd=(s/s.cummax()-1).min(); ww=[(s.loc[a:b]/s.loc[a:b].cummax()-1).min()*100 for a,b in WIN]
    i=i0; w=0; tot=0
    while i<n:
        sg=1 if E[i]>0 else (-1 if E[i]<0 else 0)
        if sg==0: i+=1; continue
        j=i
        while j+1<n and (1 if E[j+1]>0 else (-1 if E[j+1]<0 else 0))==sg: j+=1
        f0=eq[i-1] if i>0 else 500.0
        if eq[j]/f0-1>0: w+=1
        tot+=1; i=j+1
    return s.iloc[-1],cagr,sh,(cagr/abs(dd) if dd<0 else 0),dd,(w/tot if tot else 0),ww,s

# (name, signal, cap, vt, smooth, band, gates, legacy, chart?)
M=[("Buy & Hold (BTC)", np.ones(n),1,999,0,0.0,False,None,True),
   ("Raw 8B 5x (original)", expf,5,0.6,0,0.0,True,5,False),
   ("Core 1x (PRESERVE)", expf,1,0.6,5,0.15,True,None,True),
   ("Balanced 2x (cap2)", expf,2,1.5,5,0.15,True,None,False),
   ("Balanced 2.5x (cap2.5, dead-zone)", expf,2.5,1.5,5,0.15,True,None,False),
   ("Balanced 3x (cap3) *REC", expf,3,1.5,5,0.15,True,None,True),
   ("Growth 5x = A", expf,5,1.5,5,0.15,True,None,True),
   ("Aggressive B (vt2/cap5)", expf,5,2.0,5,0.25,True,None,False),
   ("Smooth C (sm10/cap5)", expf,5,1.5,10,0.25,True,None,False),
   ("Conditional-lev", expf,condcap,1.5,5,0.15,True,None,False),
   ("Cycle math (in-sample)", cyc,5,0.8,5,0.15,False,None,True)]
res=[]
print(f"{'MODEL':28s} {'$@0bp':>14s} {'$@50bp':>13s} {'$@1x':>10s} {'CAGR':>5s} {'Calm':>5s} {'Shrp':>5s} {'maxDD':>6s} {'Win%':>5s} | W1/W2/W3")
for nm,sig,cap,vt,sm,bd,gt,lg,ch in M:
    eq0,_,_=sim(sig,cap,vt,sm,bd,0.0,gt,0.30,lg)
    eq5,E5,liq=sim(sig,cap,vt,sm,bd,0.005,gt,0.30,lg)
    eq1,_,_=sim(sig,1,vt,sm,bd,0.005,gt,0.30,(1 if lg else None))
    f0=eq0[-1]; f5,cagr,sh,cal,dd,wr,ww,s5=metr(eq5,E5); f1=eq1[-1]
    res.append((nm,f0,f5,f1,cagr,cal,sh,dd,wr,ww,liq,ch,s5))
    print(f"{nm:28s} ${f0:>13,.0f} ${f5:>12,.0f} ${f1:>9,.0f} {cagr*100:>4.0f}% {cal:>5.2f} {sh:>5.2f} {dd*100:>5.0f}% {wr*100:>4.0f}% | {ww[0]:>3.0f}/{ww[1]:>3.0f}/{ww[2]:>3.0f}%")

# ===== Excel =====
HDR=Font(bold=True,color="FFFFFF"); HF=PatternFill("solid",fgColor="1F3864"); TIT=Font(bold=True,size=13)
thin=Side(style="thin",color="DDDDDD"); BD=Border(thin,thin,thin,thin)
wb=Workbook(); ws=wb.active; ws.title="Compare"
ws["A1"]="BTC — ALL models, optimized (size-controlled), full comparison"; ws["A1"].font=TIT
ws["A2"]="$500 start, 2014→now. @50bp = realistic. Calmar/Sharpe/DD/Win% at 50bp. W1=2018, W2=2021-22, W3=2025 drops."; ws["A2"].font=Font(italic=True,color="666666")
cols=["Model","$@0bp","$@50bp","$@1x","CAGR","Calmar","Sharpe","maxDD","Win%","Liq","W1 2018","W2 21-22","W3 2025"]
for j,c in enumerate(cols,1):
    cc=ws.cell(4,j,c); cc.font=HDR; cc.fill=HF; cc.alignment=Alignment(horizontal="center",wrap_text=True)
for k,(nm,f0,f5,f1,cagr,cal,sh,dd,wr,ww,liq,ch,s5) in enumerate(res):
    r=5+k
    vals=[nm,round(f0),round(f5),round(f1),cagr,cal,sh,dd,wr,liq,ww[0]/100,ww[1]/100,ww[2]/100]
    for j,v in enumerate(vals,1):
        cc=ws.cell(r,j,v); cc.border=BD
    for j in (2,3,4): ws.cell(r,j).number_format='#,##0'
    for j in (5,9,11,12,13): ws.cell(r,j).number_format='0%'
    for j in (6,7): ws.cell(r,j).number_format='0.00'
    ws.cell(r,8).number_format='0%'
    if "*REC" in nm:
        for j in range(1,14): ws.cell(r,j).fill=PatternFill("solid",fgColor="E2EFDA")
ws.column_dimensions["A"].width=28
for c in "BCDEFGHIJKLM": ws.column_dimensions[c].width=11

# Calmar bar chart
bc=BarChart(); bc.title="Calmar (risk-adjusted return) by model"; bc.type="bar"; bc.height=10; bc.width=20
data=Reference(ws,min_col=6,min_row=4,max_row=4+len(res)); cats=Reference(ws,min_col=1,min_row=5,max_row=4+len(res))
bc.add_data(data,titles_from_data=True); bc.set_categories(cats); ws.add_chart(bc,"O4")

# Equity curves sheet
we=wb.create_sheet("Equity Curves")
chart_models=[(nm,s5) for (nm,f0,f5,f1,cagr,cal,sh,dd,wr,ww,liq,ch,s5) in res if ch]
ddt=[d.strftime("%Y-%m-%d") for d in chart_models[0][1].index]
we.cell(1,1,"Date")
for j,(nm,s5) in enumerate(chart_models,2): we.cell(1,j,nm)
for i,d in enumerate(ddt,2):
    we.cell(i,1,d)
    for j,(nm,s5) in enumerate(chart_models,2): we.cell(i,j,round(float(s5.iloc[i-2]),2))
for cc in we[1]: cc.font=Font(bold=True)
last=len(ddt)+1
lc=LineChart(); lc.title="Equity $500→ @50bp (log scale)"; lc.y_axis.title="Equity $"; lc.x_axis.title="Date"
lc.height=12; lc.width=26; lc.y_axis.scaling.logBase=10
for j in range(2,2+len(chart_models)):
    lc.add_data(Reference(we,min_col=j,min_row=1,max_row=last),titles_from_data=True)
lc.set_categories(Reference(we,min_col=1,min_row=2,max_row=last)); we.add_chart(lc, get_column_letter(len(chart_models)+3)+"2")

# Rejected sheet
wr_=wb.create_sheet("Rejected approaches")
wr_["A1"]="Approaches tested and REJECTED (no robust edge → can't be optimized)"; wr_["A1"].font=TIT
rej=[["Approach","Result / why rejected"],
 ["Cut-loss (fixed %)","Whipsaw; every level worse @50bp; -doesn't fix slippage"],
 ["Cut-loss (trailing)","Turnover 34→90x; @50bp B2x $179k→$1; -100% DD"],
 ["Discrete hold + stop","Same whipsaw; ruin at 50bp"],
 ["Tail hedge (OTM puts 30/90d)","Premium ~3x payoff (vol-risk-premium); cuts return, no DD help"],
 ["MVRV cycle-top overlay","Separability AUC 0.36-0.56; overlay cut B2x 362x→36x, fails OOS"],
 ["Fear&Greed contrarian","Cuts longs in greed = fights trend; -70% return"],
 ["Funding-only gate","Inconsistent (helps 2x, hurts 3x); 2020+ only"],
 ["Power-law mean-reversion","Sign is momentum not MR; MR backtest → ~$0"],
 ["Halving-cycle as signal","Train Sh 0.27 / test 1.22 = luck on 3-4 cycles (not robust)"],
 ["Actuary kitchen-sink regression","OOS IC -0.01 (worse than coinflip); ensemble IC +0.15"],
 ["Reversal/exhaustion prediction","Composite AUC 0.61; rule hurts & fails OOS"],
 ["Order-flow / depth","Standalone IC +0.02 (sub-daily signal; no daily edge)"]]
for i,row in enumerate(rej,3):
    for j,v in enumerate(row,1):
        cc=wr_.cell(i,j,v)
        if i==3: cc.font=HDR; cc.fill=HF
wr_.column_dimensions["A"].width=30; wr_.column_dimensions["B"].width=70

wb.save(OUT); print("\nsaved", OUT, "| sheets:", wb.sheetnames)
