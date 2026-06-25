"""Comprehensive comparison workbook (with an embedded return chart) for the NEW
finer-regime system. Scenarios run in lockstep (same trades, different leverage):
  - SPOT 1x (confidence-sized)
  - LEV 5x long / 2x short (confidence-sized, HONEST gap-liquidation)
Plus the diversified ensemble curve and Buy&Hold for comparison on the chart.

Output: New setup for BTC/BTC_Power_compare.xlsx
Sheets: Summary | Equity (+ native line chart) | Trades | By Year
"""
import os
import numpy as np
import pandas as pd
import compare_m1m5 as cm
import regime_v2 as r2
import signals as sg
import backtest as bt
import stable_combo as sc
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Font, PatternFill, Alignment

HERE = os.path.dirname(__file__)
OUTX = os.path.join(HERE, "..", "..", "BTC_Power_compare.xlsx")
ANN = 365
FEE = 0.0005
ENGINES = list(sg.EXCEL_COL.keys())


def best_per_regime(df, reg, memb, floor=0.5):
    close = df["close"].to_numpy(); px = np.zeros(len(df)); px[1:] = close[1:] / close[:-1] - 1
    rprev = np.roll(reg, 1); mid = len(df) // 2; out = {}
    for cell in r2.CELLS:
        mask = rprev == cell; best = None
        for k in ENGINES:
            held = np.roll(memb[k], 1); r = held * px
            m1 = mask.copy(); m1[mid:] = False; m2 = mask.copy(); m2[:mid] = False
            if r[m1].std() > 0 and r[m2].std() > 0 and m1.sum() > 12 and m2.sum() > 12:
                s = min(r[m1].mean() * ANN / (r[m1].std(ddof=1) * np.sqrt(ANN)),
                        r[m2].mean() * ANN / (r[m2].std(ddof=1) * np.sqrt(ANN)))
                if best is None or s > best[1]:
                    best = (k, s)
        out[cell] = best if (best and best[1] >= floor) else None
    return out


def conf_bucket(score):
    return ("High", 1.0) if score >= 1.5 else (("Med", 0.7) if score >= 0.8 else ("Low", 0.4))


def run(df, reg, memb, bpr):
    """Lockstep discrete engine: SPOT (1x) + LEV (5x long/2x short), conf-sized, MTM daily,
    honest gap-liquidation for the leveraged book."""
    close = df["close"].to_numpy(); low = df["low"].to_numpy(); high = df["high"].to_numpy()
    dates = df["Date"].tolist(); n = len(df); i0 = 260
    eqs = 500.0; eql = 500.0; pos = None
    es = np.full(n, 500.0); el = np.full(n, 500.0)
    trades = []; liql = 0
    for i in range(i0, n):
        rg = reg[i]
        if pos is not None:
            d = pos["dir"]; cl = pos["cl"]; ex = False; fillp = None; reason = None; liq = False
            # honest leverage liquidation (gap past stop): long liq at entry*(1-1/2lev?) use 1/lev
            liqpx_l = pos["entry"] * (1 - 1 / 5.0 + 0.01) if d == 1 else pos["entry"] * (1 + 1 / 2.0 - 0.01)
            if d == 1:
                if low[i] <= liqpx_l and low[i] <= pos["stop"]:   # gapped through stop into liq
                    liq = True; fillp = liqpx_l; reason = "LIQUIDATED (5x gap)"; ex = True
                elif low[i] <= pos["stop"]:
                    fillp = pos["stop"]; reason = "Trailing stop"; ex = True
                elif high[i] > pos["hi"]:
                    pos["hi"] = high[i]; pos["stop"] = max(pos["stop"], pos["hi"] * (1 - cl))
            else:
                if high[i] >= liqpx_l and high[i] >= pos["stop"]:
                    liq = True; fillp = liqpx_l; reason = "LIQUIDATED (2x gap)"; ex = True
                elif high[i] >= pos["stop"]:
                    fillp = pos["stop"]; reason = "Trailing stop"; ex = True
                elif low[i] < pos["lo"]:
                    pos["lo"] = low[i]; pos["stop"] = min(pos["stop"], pos["lo"] * (1 + cl))
            if not ex:
                m = memb[pos["eng"]][i]
                su = pos["regime"] == rg
                sigx = (m * d <= 0)
                if (not su) or sigx:
                    fillp = close[i]; reason = "Regime/Signal exit"; ex = True
            if ex:
                ret = (fillp / pos["entry"] - 1.0) * d
                pnl_s = pos["ns"] * ret; eqs += pnl_s - pos["ns"] * FEE * 2
                if liq:
                    pnl_l = -pos["ml"]; eql += pnl_l - pos["nl"] * FEE; liql += 1   # lose the margin
                else:
                    pnl_l = pos["nl"] * ret; eql += pnl_l - pos["nl"] * FEE * 2
                trades.append(dict(entry_dt=pos["edt"], exit_dt=dates[i], market=pos["regime"],
                                   engine=pos["eng"], direction="LONG" if d == 1 else "SHORT",
                                   confidence=pos["bk"], entry=round(pos["entry"], 2),
                                   exit=round(float(fillp), 2), ret=round(ret, 4),
                                   cutloss=round(pos["cutpx"], 2), sizing=pos["dep"],
                                   lev=(5 if d == 1 else 2), pnl_spot=round(pnl_s, 2),
                                   pnl_lev=round(pnl_l, 2), reason=reason,
                                   eq_spot=round(eqs, 2), eq_lev=round(max(eql, 0), 2)))
                pos = None
        if pos is None and i >= i0 and eqs > 1:
            be = bpr.get(rg)
            if be is not None:
                m = memb[be[0]][i]; d = 1 if m > 0 else (-1 if m < 0 else 0)
                if d != 0:
                    bk, dep = conf_bucket(be[1]); cl = 0.10 if d == 1 else 0.07
                    entry = close[i]; ns = dep * eqs; ml = dep * max(eql, 0)
                    nl = ml * (5 if d == 1 else 2)
                    cutpx = entry * (1 - cl) if d == 1 else entry * (1 + cl)
                    pos = dict(entry=entry, edt=dates[i], dir=d, eng=be[0], regime=rg, bk=bk, dep=dep,
                               cl=cl, hi=entry, lo=entry, stop=cutpx, cutpx=cutpx, ns=ns, nl=nl, ml=ml)
        es[i] = eqs; el[i] = max(eql, 0)
    return es, el, trades, liql


def met(eq):
    eq = np.asarray(eq, float); r = np.diff(eq) / np.where(eq[:-1] == 0, 1, eq[:-1]); r = r[np.isfinite(r)]
    yrs = len(eq) / ANN; cagr = (eq[-1] / 500) ** (1 / yrs) - 1 if eq[-1] > 0 else -1
    sh = r.mean() * ANN / (r.std(ddof=1) * np.sqrt(ANN)) if r.std() > 0 else float("nan")
    dd = (eq / np.maximum.accumulate(eq) - 1).min()
    return dict(final=eq[-1], cagr=cagr, sharpe=sh, maxdd=dd, calmar=cagr / abs(dd) if dd < 0 else float("nan"))


def main():
    df, reg, memb = sc.prep()
    bpr = best_per_regime(df, reg, memb)
    es, el, trades, liql = run(df, reg, memb, bpr)
    # diversified-ensemble scenarios (the recommended system) for the chart/summary
    emap = sc.eligible_map(df, reg, memb); exp = sc.exposure_series(df, reg, memb, emap)
    close = df["close"].to_numpy(); i0 = 260
    bh = np.full(len(df), 500.0); bh[i0:] = 500.0 * close[i0:] / close[i0]
    LOOSE = dict(vol_target=0.60, dd_kill=0.30)               # max growth
    TIGHT = dict(vol_target=0.40, dd_derisk=0.40, smooth=3)    # smoother / min drawdown
    scen = {}
    for name, lev, ctrl in [("Ensemble 1x (recommended)", 1, LOOSE),
                            ("Ensemble 1x smoother", 1, TIGHT),
                            ("Ensemble 3x growth", 3, LOOSE),
                            ("Ensemble 5x growth", 5, LOOSE),
                            ("Ensemble 5x lower-DD", 5, TIGHT)]:
        eq, lq = sc.simulate(df, exp, lev=lev, slip=0.003, **ctrl)
        scen[name] = (eq, lq)
    scen["Buy & Hold"] = (bh, 0)
    scen["SPOT 1x single-engine"] = (es, 0)
    scen["LEV 5x/2x naive (RUIN)"] = (el, liql)
    dates = df["Date"].tolist()
    print("best engine per regime:", {k: (v[0] if v else None) for k, v in bpr.items()})
    for nm, (eq, lq) in scen.items():
        m = met(eq); print(f"  {nm:30s} ${m['final']:>13,.0f} DD{m['maxdd']*100:>6.0f}% Sh{m['sharpe']:>5.2f} Cal{m['calmar']:>5.2f} liq{lq}")
    write_xlsx(df, dates, scen, trades, bpr)
    print("->", OUTX)


CHART_SCEN = ["Ensemble 1x (recommended)", "Ensemble 1x smoother", "Ensemble 3x growth",
              "Ensemble 5x growth", "Ensemble 5x lower-DD", "Buy & Hold"]


def write_xlsx(df, dates, scen, trades, bpr):
    i0 = 260
    d = dates[i0:]
    eqcols = {"Date": d}
    for name in CHART_SCEN:
        eqcols[name] = np.round(scen[name][0][i0:], 2)
    eqdf = pd.DataFrame(eqcols)
    # scenarios summary (all scenarios, incl. reference rows)
    order = CHART_SCEN[:-1] + ["Buy & Hold", "SPOT 1x single-engine", "LEV 5x/2x naive (RUIN)"]
    rows = []
    for name in order:
        eq, lq = scen[name]; m = met(eq)
        rows.append(dict(Scenario=name, **{"$500 ->": round(m["final"], 0), "CAGR": f"{m['cagr']*100:.0f}%",
                    "Sharpe": round(m["sharpe"], 2), "maxDD": f"{m['maxdd']*100:.0f}%",
                    "Calmar": round(m["calmar"], 2), "Liquidations": lq}))
    sumdf = pd.DataFrame(rows)
    # trades
    tdf = pd.DataFrame(trades); tdf.insert(0, "#", range(1, len(tdf) + 1))
    wins = (tdf["ret"] > 0).sum()
    notes = pd.DataFrame({"Scenario": ["", "Win ratio (overall)", "Win LONG", "Win SHORT",
                                       "Trades", "Best engine per regime", "",
                                       "SPOT 1x = no leverage, no liquidation (recommended).",
                                       "LEV 5x/2x = confidence-sized; gap-liquidation modeled honestly.",
                                       "Hypothetical 2014+; fees+funding+slippage. Not advice."],
                          "$500 ->": ["", f"{wins/len(tdf)*100:.1f}%",
                                      f"{((tdf['direction']=='LONG')&(tdf['ret']>0)).sum()}/{(tdf['direction']=='LONG').sum()}",
                                      f"{((tdf['direction']=='SHORT')&(tdf['ret']>0)).sum()}/{(tdf['direction']=='SHORT').sum()}",
                                      len(tdf), str({k: (v[0] if v else 'aside') for k, v in bpr.items()}), "", "", "", ""]})
    sumdf = pd.concat([sumdf, notes], ignore_index=True)
    # by year
    tdf["_y"] = tdf["exit_dt"].str[:4]; yr = []
    for y, g in tdf.groupby("_y"):
        w = (g["ret"] > 0).sum()
        yr.append(dict(Year=y, Trades=len(g), **{"Win %": f"{w/len(g)*100:.1f}%",
                  "PnL spot $": round(g["pnl_spot"].sum(), 0), "PnL lev $": round(g["pnl_lev"].sum(), 0)}))
    ydf = pd.DataFrame(yr); tdf = tdf.drop(columns=["_y"])

    with pd.ExcelWriter(OUTX, engine="openpyxl") as xw:
        sumdf.to_excel(xw, sheet_name="Summary", index=False)
        eqdf.to_excel(xw, sheet_name="Equity", index=False)
        tdf.to_excel(xw, sheet_name="Trades", index=False)
        ydf.to_excel(xw, sheet_name="By Year", index=False)
        wb = xw.book
        # chart on Equity sheet
        ws = wb["Equity"]; nrow = len(eqdf) + 1
        ch = LineChart(); ch.title = "Equity from $500 (log scale)"; ch.height = 11; ch.width = 26
        ch.y_axis.scaling.logBase = 10; ch.y_axis.title = "Equity $ (log)"; ch.x_axis.title = "Date"
        ch.x_axis.delete = False; ch.y_axis.delete = False
        data = Reference(ws, min_col=2, max_col=1 + len(CHART_SCEN), min_row=1, max_row=nrow)
        cats = Reference(ws, min_col=1, min_row=2, max_row=nrow)
        ch.add_data(data, titles_from_data=True); ch.set_categories(cats)
        for s in ch.series:
            s.graphicalProperties.line.width = 18000
        ws.add_chart(ch, "G2")
        _fmt(wb)
    return


def _fmt(wb):
    hf = PatternFill("solid", fgColor="1F2A44"); ft = Font(bold=True, color="FFFFFF")
    for nm in wb.sheetnames:
        ws = wb[nm]; ws.freeze_panes = "A2"
        for c in ws[1]:
            c.fill = hf; c.font = ft; c.alignment = Alignment(horizontal="center")
        for col in ws.columns:
            try:
                w = max(len(str(c.value)) if c.value is not None else 0 for c in col)
                ws.column_dimensions[col[0].column_letter].width = min(max(w + 2, 10), 24)
            except Exception:
                pass
    ws = wb["Trades"]; hdr = [c.value for c in ws[1]]
    fmt = {"ret": "0.00%", "sizing": "0%", "entry": "#,##0", "exit": "#,##0", "cutloss": "#,##0",
           "pnl_spot": "#,##0;[Red]-#,##0", "pnl_lev": "#,##0;[Red]-#,##0", "eq_spot": "#,##0", "eq_lev": "#,##0"}
    for j, h in enumerate(hdr, 1):
        if h in fmt:
            for r in range(2, ws.max_row + 1):
                ws.cell(row=r, column=j).number_format = fmt[h]


if __name__ == "__main__":
    main()
