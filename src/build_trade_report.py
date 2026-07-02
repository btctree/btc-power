"""Full comparative trade report for 3 size-control packages, across 0bp / 50bp / 1x.
A = vt1.5 cap5 sm5 bd0.15 | B = vt2.0 cap5 sm5 bd0.25 | C = vt1.5 cap5 sm10 bd0.25
Per trade (direction-spell): entry/exit date+price+why, days, regime/engines, funds deployed,
avg exposure, notional, margin used + %, P&L $ and % (50bp), and P&L% at 0bp & 1x. Plus summary
with win%, avg win/loss, max consecutive wins/losses. Honest: 2014+, VOL+FUND gates, dd-kill, liquidation.
"""
import os
import numpy as np, pandas as pd
import stable_combo as sc
import live_engine as le
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

ANN = 365
OUT = r"C:\Users\user\OneDrive\Desktop\New setup for BTC\BTC_trade_report.xlsx"
df, reg0, memb = sc.prep()
reg, emap, exp_raw = le.ensemble_ctx(df, memb)
expf = np.where(np.abs(exp_raw) >= 0.4, exp_raw, 0.0)
close = df["close"].to_numpy(); low = df["low"].to_numpy(); high = df["high"].to_numpy()
n = len(df); i0 = 260; dstr = df["Date"].tolist(); dates = pd.to_datetime(df["Date"])
rv = pd.Series(close).pct_change().rolling(20).std().to_numpy() * np.sqrt(ANN)
HERE = os.path.dirname(os.path.abspath(__file__))
def lmap(p, col):
    if not os.path.exists(p): return {}
    f = pd.read_csv(p); return dict(zip(f.iloc[:, 0].astype(str), f[col]))
funding = np.array([lmap(os.path.join(HERE, "..", "data", "funding.csv"), "funding_rate").get(d, np.nan) for d in dstr])
def trail_rank(a, w=365):
    return pd.Series(a).rolling(w, min_periods=60).apply(lambda x: (x.iloc[-1] >= x).mean(), raw=False).to_numpy()
vol_rank = trail_rank(rv); fund_rank = trail_rank(funding)
gl = np.ones(n); gs = np.ones(n)
for i in range(n):
    if vol_rank[i] == vol_rank[i] and vol_rank[i] > 0.85: gl[i] *= 0.5; gs[i] *= 0.5
    if fund_rank[i] == fund_rank[i]:
        if fund_rank[i] > 0.90: gl[i] *= 0.5
        if fund_rank[i] < 0.10: gs[i] *= 0.5

def run(vt, cap, smooth, band, slip, dd_kill=0.30, fee=0.0005, maint=0.01):
    e_in = pd.Series(expf).ewm(span=smooth, adjust=False).mean().to_numpy() if smooth > 1 else expf.copy()
    equity = peak = 500.0; held = 0.0; eq = np.full(n, 500.0); E = np.zeros(n)
    for i in range(i0, n):
        sig = e_in[i - 1]; g = gl[i - 1] if sig > 0 else (gs[i - 1] if sig < 0 else 1.0)
        if not (rv[i - 1] == rv[i - 1] and rv[i - 1] > 0): eq[i] = equity; E[i] = held; continue
        e = sig * g * min(cap, vt / rv[i - 1])
        if dd_kill > 0 and equity < peak * (1 - dd_kill): e *= 0.5
        if band > 0 and abs(e - held) < band and not (e == 0 and held != 0): e = held
        adv = (-(low[i] / close[i - 1] - 1)) if e > 0 else ((high[i] / close[i - 1] - 1) if e < 0 else 0.0)
        if e != 0 and abs(e) * max(adv, 0) >= (1 - maint):
            equity *= 0.01; held = 0.0; eq[i] = equity; E[i] = 0.0; peak = max(peak, equity); continue
        equity *= (1 + e * (close[i] / close[i - 1] - 1)); equity -= equity * abs(e - held) * (fee + slip)
        held = e; eq[i] = max(equity, 1e-9); E[i] = e; peak = max(peak, equity)
    return eq, E

def spells(E):
    """direction spells from signed exposure E (scenario-independent direction)."""
    out = []; i = i0
    while i < n:
        s = 1 if E[i] > 0 else (-1 if E[i] < 0 else 0)
        if s == 0: i += 1; continue
        j = i
        while j + 1 < n and (1 if E[j + 1] > 0 else (-1 if E[j + 1] < 0 else 0)) == s: j += 1
        nxt = (1 if (j + 1 < n and E[j + 1] > 0) else (-1 if (j + 1 < n and E[j + 1] < 0) else 0))
        out.append((i, j, s, nxt)); i = j + 1
    return out

def metrics(eq):
    s = pd.Series(eq[i0:]); r = s.pct_change().dropna(); yrs = len(s) / ANN
    cagr = (s.iloc[-1] / 500) ** (1 / yrs) - 1 if s.iloc[-1] > 0 else -1
    sh = r.mean() * ANN / (r.std(ddof=1) * np.sqrt(ANN)) if r.std() > 0 else float("nan")
    dd = (s / s.cummax() - 1).min()
    return s.iloc[-1], cagr, sh, (cagr / abs(dd) if dd < 0 else float("nan")), dd

CONFIGS = [("A vt1.5/cap5/sm5/bd.15", 1.5, 5, 5, 0.15),
           ("B vt2.0/cap5/sm5/bd.25", 2.0, 5, 5, 0.25),
           ("C vt1.5/cap5/sm10/bd.25", 1.5, 5, 10, 0.25)]
SCEN = [("0bp-lev", "cap", 0.0), ("50bp-lev", "cap", 0.005), ("1x", 1.0, 0.005)]

wb = Workbook(); HDR = Font(bold=True, color="FFFFFF"); HF = PatternFill("solid", fgColor="1F3864")
TIT = Font(bold=True, size=13); thin = Side(style="thin", color="DDDDDD"); BD = Border(thin, thin, thin, thin)

# ---------- Summary ----------
ws = wb.active; ws.title = "Summary"
ws["A1"] = "BTC size-control packages — trade report (A/B/C across 0bp / 50bp / 1x)"; ws["A1"].font = TIT
cols = ["Config", "Scenario", "Final $", "CAGR", "Sharpe", "Calmar", "maxDD", "Trades", "Win %",
        "Avg win %", "Avg loss %", "Max consec win", "Max consec loss", "Avg exposure", "Avg margin %"]
for j, cc in enumerate(cols, 1):
    c = ws.cell(3, j, cc); c.font = HDR; c.fill = HF; c.alignment = Alignment(horizontal="center", wrap_text=True)
rr = 3
trade_cache = {}
for name, vt, cap, sm, bd in CONFIGS:
    for sname, scap, slip in SCEN:
        ucap = cap if scap == "cap" else scap
        eq, E = run(vt, ucap, sm, bd, slip)
        sp = spells(E)
        rows = []
        for (s, t, d, nxt) in sp:
            f0 = eq[s - 1] if s > 0 else 500.0
            pnl = eq[t] / f0 - 1 if f0 > 0 else 0.0
            avgexp = float(np.mean(np.abs(E[s:t + 1])))
            rows.append(dict(s=s, t=t, d=d, nxt=nxt, funds=f0, pnl=pnl, avgexp=avgexp,
                             notional=avgexp * f0, marginpct=avgexp / ucap))
        trade_cache[(name, sname)] = rows
        pnls = np.array([r["pnl"] for r in rows]); wins = pnls > 0
        # consecutive
        mcw = mcl = cw = cl = 0
        for w in wins:
            if w: cw += 1; cl = 0; mcw = max(mcw, cw)
            else: cl += 1; cw = 0; mcl = max(mcl, cl)
        fin, cagr, sh, cal, dd = metrics(eq)
        vals = [name, sname, round(fin), cagr, sh, cal, dd, len(rows),
                wins.mean() if len(rows) else 0,
                pnls[wins].mean() if wins.any() else 0,
                pnls[~wins].mean() if (~wins).any() else 0, mcw, mcl,
                np.mean([r["avgexp"] for r in rows]) if rows else 0,
                np.mean([r["marginpct"] for r in rows]) if rows else 0]
        rr += 1
        for j, v in enumerate(vals, 1):
            cc = ws.cell(rr, j, v); cc.border = BD
        ws.cell(rr, 3).number_format = '#,##0'
        for j in (4, 9, 10, 11, 15): ws.cell(rr, j).number_format = '0.0%'
        ws.cell(rr, 7).number_format = '0%'
        for j in (5, 6, 14): ws.cell(rr, j).number_format = '0.00'
    rr += 1
ws.column_dimensions["A"].width = 22; ws.column_dimensions["B"].width = 10
for col in "CDEFGHIJKLMNO": ws.column_dimensions[col].width = 11

# ---------- per-config trade logs (detail = 50bp-lev; + P&L% at 0bp & 1x) ----------
for name, vt, cap, sm, bd in CONFIGS:
    wsx = wb.create_sheet(name.split()[0] + " trades")
    r50 = trade_cache[(name, "50bp-lev")]; r0 = {r["s"]: r for r in trade_cache[(name, "0bp-lev")]}
    r1 = {r["s"]: r for r in trade_cache[(name, "1x")]}
    hd = ["#", "Entry date", "Entry $", "Dir", "Market", "Engines", "Why IN", "Exit date", "Exit $",
          "Days", "Why OUT", "Funds @entry $", "Avg exp (x)", "Notional $", "Margin %",
          "P&L $ (50bp)", "P&L % (50bp)", "P&L % (0bp)", "P&L % (1x)"]
    for j, h in enumerate(hd, 1):
        c = wsx.cell(1, j, h); c.font = HDR; c.fill = HF; c.alignment = Alignment(wrap_text=True)
    for k, r in enumerate(r50, 1):
        s, t, d, nxt = r["s"], r["t"], r["d"], r["nxt"]
        rgm = reg[s]; eng = ", ".join(emap.get(rgm) or []) or "-"
        whyin = f"ensemble turned {'LONG' if d > 0 else 'SHORT'} ({rgm})"
        whyout = ("flipped to " + ("LONG" if nxt > 0 else "SHORT")) if nxt != 0 else "went flat (conviction/dd-kill)"
        pnl0 = r0.get(s, {}).get("pnl", float("nan")); pnl1 = r1.get(s, {}).get("pnl", float("nan"))
        pnl_usd = r["funds"] * r["pnl"]
        row = [k, dstr[s], round(close[s], 2), "LONG" if d > 0 else "SHORT", rgm, eng, whyin,
               dstr[t], round(close[t], 2), t - s + 1, whyout, round(r["funds"], 2),
               round(r["avgexp"], 2), round(r["notional"], 2), r["marginpct"],
               round(pnl_usd, 2), r["pnl"], pnl0, pnl1]
        for j, v in enumerate(row, 1):
            wsx.cell(k + 1, j, v)
        for j in (15, 17, 18, 19): wsx.cell(k + 1, j).number_format = '0.0%'
        for j in (3, 9, 12, 14, 16): wsx.cell(k + 1, j).number_format = '#,##0'
    for col, w in [("B", 11), ("E", 12), ("F", 16), ("G", 24), ("H", 11), ("K", 22)]:
        wsx.column_dimensions[col].width = w

wb.save(OUT)
print("saved", OUT)
for name, _, _, _, _ in CONFIGS:
    for sn in ["0bp-lev", "50bp-lev", "1x"]:
        rows = trade_cache[(name, sn)]
        print(f"{name:26s} {sn:9s} trades {len(rows)}")
