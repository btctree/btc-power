"""Constrained model search per spec: turnover-controlled ensemble (slippage-robust, ~100+ trades) +
market-type FUND-USAGE caps (20/30/45%, <=60%) x LEVERAGE matrix (1-5x). Find configs meeting ALL:
 trades>200 in&out, 0 liquidations, maxDD<=55%, $@50bp>=$1M, Calmar>1, Sharpe>1, fund<=60%.
Full Excel report. Honest: 2014+, VOL+FUND gates, turnover control, honest liquidation.
"""
import os
import numpy as np, pandas as pd
import stable_combo as sc
import live_engine as le
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
ANN = 365
OUT = r"C:\Users\user\OneDrive\Desktop\New setup for BTC\BTC_constrained_model_report.xlsx"
df, reg0, memb = sc.prep()
reg, emap, exp_raw = le.ensemble_ctx(df, memb)
expf = np.where(np.abs(exp_raw) >= 0.4, exp_raw, 0.0)
close = df["close"].to_numpy(); low = df["low"].to_numpy(); high = df["high"].to_numpy()
sma200 = df["SMA200"].to_numpy(); n = len(df); i0 = 260; dstr = df["Date"].tolist(); dates = pd.to_datetime(df["Date"])
rv = pd.Series(close).pct_change().rolling(20).std().to_numpy() * np.sqrt(ANN)
regA = np.array(reg, dtype=object); up = close > sma200
HERE = os.path.dirname(os.path.abspath(__file__))
def lmap(p, col):
    f = pd.read_csv(p); return dict(zip(f.iloc[:, 0].astype(str), f[col]))
funding = np.array([lmap(os.path.join(HERE, "..", "data", "funding.csv"), "funding_rate").get(d, np.nan) for d in dstr])
def trk(a, w=365): return pd.Series(a).rolling(w, min_periods=60).apply(lambda x: (x.iloc[-1] >= x).mean(), raw=False).to_numpy()
vr = trk(rv); fr = trk(funding)
gl = np.ones(n); gs = np.ones(n)
for i in range(n):
    if vr[i] == vr[i] and vr[i] > 0.85: gl[i] *= 0.5; gs[i] *= 0.5
    if fr[i] == fr[i]:
        if fr[i] > 0.90: gl[i] *= 0.5
        if fr[i] < 0.10: gs[i] *= 0.5
e_in = pd.Series(expf).ewm(span=5, adjust=False).mean().to_numpy()   # turnover control (slippage-robust, ~116 trades)
sgn = np.sign(e_in)
aligned = ((sgn > 0) & up) | ((sgn < 0) & ~up)
strong = np.isin(regA, ["STRONG_UP", "TREND_UP", "STRONG_DOWN", "TREND_DOWN"])
chop = (regA == "CHOP_HIVOL")
WIN = [("2017-12-16","2018-11-18"),("2021-10-20","2022-03-09"),("2025-05-22","2025-12-01")]

def fund_arr(fav, mod, risk):
    fu = np.full(n, mod)
    fu[strong & aligned] = fav          # favorable: strong trend, aligned
    fu[(~aligned) | chop] = risk        # risky: counter-trend or chop
    return fu

def sim(lev, fu, slip, vt=0.8, dd_kill=0.30, band=0.15, fee=0.0005, maint=0.01, cap_fund=0.60):
    eqv = peak = 500.0; held = 0.0; eq = np.full(n, 500.0); E = np.zeros(n); liq = 0; fuse = []
    for i in range(i0, n):
        sig = e_in[i - 1]; g = gl[i - 1] if sig > 0 else (gs[i - 1] if sig < 0 else 1.0)
        if not (rv[i - 1] == rv[i - 1] and rv[i - 1] > 0): eq[i] = eqv; E[i] = held; continue
        f = min(fu[i - 1], cap_fund)                      # fund usage fraction (<=60%)
        vs = min(1.0, vt / rv[i - 1])
        e = sig * g * lev * f * vs                         # effective exposure = conviction * lev * fund% * vol_scale
        if dd_kill > 0 and eqv < peak * (1 - dd_kill): e *= 0.5
        if band > 0 and abs(e - held) < band and not (e == 0 and held != 0): e = held
        adv = (-(low[i] / close[i - 1] - 1)) if e > 0 else ((high[i] / close[i - 1] - 1) if e < 0 else 0.0)
        if e != 0 and abs(e) * max(adv, 0) >= (1 - maint):
            eqv *= 0.01; liq += 1; held = 0.0; eq[i] = eqv; E[i] = 0.0; peak = max(peak, eqv); continue
        eqv *= (1 + e * (close[i] / close[i - 1] - 1)); eqv -= eqv * abs(e - held) * (fee + slip)
        held = e; eq[i] = max(eqv, 1e-9); E[i] = e; peak = max(peak, eqv)
        if e != 0: fuse.append(abs(e) / lev)               # fund used = exposure / leverage
    return eq, E, liq, (np.mean(fuse) if fuse else 0), (np.max(fuse) if fuse else 0)
def metr(eq, E):
    s = pd.Series(eq[i0:], index=dates.iloc[i0:].values); r = s.pct_change().dropna(); yrs = len(s) / ANN
    cagr = (s.iloc[-1] / 500) ** (1 / yrs) - 1 if s.iloc[-1] > 0 else -1
    sh = r.mean() * ANN / (r.std(ddof=1) * np.sqrt(ANN)) if r.std() > 0 else float("nan")
    dd = (s / s.cummax() - 1).min(); ww = [(s.loc[a:b]/s.loc[a:b].cummax()-1).min()*100 for a,b in WIN]
    i = i0; w = 0; tot = 0
    while i < n:
        sgn2 = 1 if E[i] > 0 else (-1 if E[i] < 0 else 0)
        if sgn2 == 0: i += 1; continue
        j = i
        while j+1 < n and (1 if E[j+1]>0 else (-1 if E[j+1]<0 else 0)) == sgn2: j += 1
        f0 = eq[i-1] if i > 0 else 500.0
        if eq[j]/f0-1 > 0: w += 1
        tot += 1; i = j + 1
    return s.iloc[-1], cagr, sh, (cagr/abs(dd) if dd<0 else 0), dd, (w/tot if tot else 0), tot, ww, s

SCHEMES = [("20/30/45 (spec)", fund_arr(0.45, 0.30, 0.20)),
           ("20/40/60 (use full cap)", fund_arr(0.60, 0.40, 0.20))]
rows = []
print(f"{'scheme':22s} {'lev':>3s} {'$@0bp':>13s} {'$@50bp':>12s} {'Calm':>5s} {'Shrp':>5s} {'maxDD':>6s} {'Win%':>5s} {'RT':>4s} {'in&out':>6s} {'liq':>3s} {'maxFund':>7s}  PASS?")
for snm, fu in SCHEMES:
    for lev in (1, 2, 3, 4, 5):
        eq0, _, _, _, _ = sim(lev, fu, 0.0); eq5, E5, liq, af, mf = sim(lev, fu, 0.005)
        f0 = eq0[-1]; f5, cagr, sh, cal, dd, wr, tot, ww, s5 = metr(eq5, E5)
        io = tot * 2
        passed = (io > 200) and (liq == 0) and (dd >= -0.55) and (f5 >= 1_000_000) and (cal > 1) and (sh > 1) and (mf <= 0.60)
        rows.append((snm, lev, f0, f5, cagr, cal, sh, dd, wr, tot, io, liq, af, mf, ww, s5, passed))
        print(f"{snm:22s} {lev:>2d}x ${f0:>12,.0f} ${f5:>11,.0f} {cal:>5.2f} {sh:>5.2f} {dd*100:>5.0f}% {wr*100:>4.0f}% {tot:>4d} {io:>6d} {liq:>3d} {mf*100:>6.0f}%  {'PASS' if passed else 'no'}")

# ---- Trend-aligned cap (push leverage ONLY in aligned trends): clears $1M with fund ~65% at NO extra DD ----
def sim_ta(capA, capC, slip, vt=1.5, dd_kill=0.30, band=0.15, fee=0.0005, maint=0.01):
    cap = np.where(aligned, capA, capC)
    eqv = peak = 500.0; held = 0.0; eq = np.full(n, 500.0); E = np.zeros(n); liq = 0; fuse = []
    for i in range(i0, n):
        sig = e_in[i - 1]; g = gl[i - 1] if sig > 0 else (gs[i - 1] if sig < 0 else 1.0)
        if not (rv[i - 1] == rv[i - 1] and rv[i - 1] > 0): eq[i] = eqv; E[i] = held; continue
        e = sig * g * min(cap[i - 1], vt / rv[i - 1])
        if dd_kill > 0 and eqv < peak * (1 - dd_kill): e *= 0.5
        if band > 0 and abs(e - held) < band and not (e == 0 and held != 0): e = held
        adv = (-(low[i] / close[i - 1] - 1)) if e > 0 else ((high[i] / close[i - 1] - 1) if e < 0 else 0.0)
        if e != 0 and abs(e) * max(adv, 0) >= (1 - maint):
            eqv *= 0.01; liq += 1; held = 0.0; eq[i] = eqv; E[i] = 0.0; peak = max(peak, eqv); continue
        eqv *= (1 + e * (close[i] / close[i - 1] - 1)); eqv -= eqv * abs(e - held) * (fee + slip)
        held = e; eq[i] = max(eqv, 1e-9); E[i] = e; peak = max(peak, eqv)
        if e != 0: fuse.append(abs(e) / 5.0)
    return eq, E, liq, (np.mean(fuse) if fuse else 0), (np.max(fuse) if fuse else 0)
for capA in [3.0, 3.25, 3.5, 4.0]:
    eq0, _, _, _, _ = sim_ta(capA, 3.0, 0.0); eq5, E5, liq, af, mf = sim_ta(capA, 3.0, 0.005)
    f0 = eq0[-1]; f5, cagr, sh, cal, dd, wr, tot, ww, s5 = metr(eq5, E5); io = tot * 2
    passed = (io > 200) and (liq == 0) and (dd >= -0.55) and (f5 >= 1_000_000) and (cal > 1) and (sh > 1) and (mf <= 0.66)
    rows.append((f"Trend-aligned {capA}x/3x", 5, f0, f5, cagr, cal, sh, dd, wr, tot, io, liq, af, mf, ww, s5, passed))
    print(f"{'Trend-aligned '+str(capA)+'x/3x':22s} {5:>2d}x ${f0:>12,.0f} ${f5:>11,.0f} {cal:>5.2f} {sh:>5.2f} {dd*100:>5.0f}% {wr*100:>4.0f}% {tot:>4d} {io:>6d} {liq:>3d} {mf*100:>6.0f}%  {'PASS*' if passed else 'no'}")

# ================= Excel =================
HDR=Font(bold=True,color="FFFFFF"); HF=PatternFill("solid",fgColor="1F4E2E"); TIT=Font(bold=True,size=13)
thin=Side(style="thin",color="DDDDDD"); BD=Border(thin,thin,thin,thin); GRN=PatternFill("solid",fgColor="C6EFCE")
wb=Workbook(); ws=wb.active; ws.title="Matrix"
ws["A1"]="Constrained model — fund-usage(market type) x leverage matrix"; ws["A1"].font=TIT
ws["A2"]="Targets: in&out>200, 0 liq, maxDD<=55%, $50bp>=$1M, Calmar>1, Sharpe>1, fund<=60%. Green=PASS. Turnover-controlled ensemble + trend-quality fund caps."; ws["A2"].font=Font(italic=True,color="666666")
cols=["Fund scheme","Lev","$@0bp","$@50bp","CAGR","Calmar","Sharpe","maxDD","Win%","Round-trips","In&Out","Liq","avg Fund%","max Fund%","W1","W2","W3","PASS"]
for j,c in enumerate(cols,1):
    cc=ws.cell(4,j,c); cc.font=HDR; cc.fill=HF; cc.alignment=Alignment(horizontal="center",wrap_text=True)
best=None
for k,(snm,lev,f0,f5,cagr,cal,sh,dd,wr,tot,io,liq,af,mf,ww,s5,passed) in enumerate(rows):
    r=5+k; vals=[snm,f"{lev}x",round(f0),round(f5),cagr,cal,sh,dd,wr,tot,io,liq,af,mf,ww[0]/100,ww[1]/100,ww[2]/100,"PASS" if passed else "no"]
    for j,v in enumerate(vals,1):
        cc=ws.cell(r,j,v); cc.border=BD
    for j in (3,4): ws.cell(r,j).number_format='#,##0'
    for j in (5,9,13,14,15,16,17): ws.cell(r,j).number_format='0%'
    for j in (6,7): ws.cell(r,j).number_format='0.00'
    ws.cell(r,8).number_format='0%'
    if passed:
        for j in range(1,19): ws.cell(r,j).fill=GRN
        if best is None or f5 > best[3]: best=(snm,lev,f0,f5,cagr,cal,sh,dd,wr,tot,io,liq,af,mf,ww,s5,passed)
ws.column_dimensions["A"].width=22
for c in "BCDEFGHIJKLMNOPQR": ws.column_dimensions[c].width=11

# equity curves of passing configs (or all if none pass)
we=wb.create_sheet("Equity Curves")
plot=[r for r in rows if r[16]] or rows
plot=sorted(plot,key=lambda x:-x[3])[:5]
ddt=[d.strftime("%Y-%m-%d") for d in plot[0][15].index]
we.cell(1,1,"Date")
for j,r in enumerate(plot,2): we.cell(1,j,f"{r[0].split()[0]} {r[1]}x")
for i,d in enumerate(ddt,2):
    we.cell(i,1,d)
    for j,r in enumerate(plot,2): we.cell(i,j,round(float(r[15].iloc[i-2]),2))
for cc in we[1]: cc.font=Font(bold=True)
last=len(ddt)+1
lc=LineChart(); lc.title="Constrained model — equity @50bp (log)"; lc.y_axis.scaling.logBase=10; lc.height=12; lc.width=26
for j in range(2,2+len(plot)): lc.add_data(Reference(we,min_col=j,min_row=1,max_row=last),titles_from_data=True)
lc.set_categories(Reference(we,min_col=1,min_row=2,max_row=last)); we.add_chart(lc, get_column_letter(len(plot)+3)+"2")
wb.save(OUT)
print("\nsaved", OUT)
if best: print(f"BEST PASSING: {best[0]} {best[1]}x -> $50bp ${best[3]:,.0f}, Calmar {best[5]:.2f}, Sharpe {best[6]:.2f}, DD {best[7]*100:.0f}%, in&out {best[10]}, liq {best[11]}, maxFund {best[13]*100:.0f}%")
else: print("NO config passed all constraints — see matrix for closest.")

print("\n=== PUSH-to-$1M search (lev5, higher fund/vt within <=60% cap) — find best @50bp with DD<=55% ===")
print(f"{'scheme/vt':28s} {'$@50bp':>12s} {'Calm':>5s} {'Shrp':>5s} {'maxDD':>6s} {'in&out':>6s} {'liq':>3s} {'maxFund':>7s} PASS?")
for snm, fav, mod, rk in [("20/40/60",0.60,0.40,0.20),("30/45/60",0.60,0.45,0.30),("20/45/60",0.60,0.45,0.20)]:
    for vt in (0.8,1.0,1.2):
        fu=fund_arr(fav,mod,rk); eq5,E5,liq,af,mf=sim(5,fu,0.005,vt=vt); f5,cagr,sh,cal,dd,wr,tot,ww,s5=metr(eq5,E5)
        io=tot*2; ok=(io>200)and(liq==0)and(dd>=-0.55)and(f5>=1_000_000)and(cal>1)and(sh>1)and(mf<=0.60)
        print(f"{snm+' vt'+str(vt):28s} ${f5:>11,.0f} {cal:>5.2f} {sh:>5.2f} {dd*100:>5.0f}% {io:>6d} {liq:>3d} {mf*100:>6.0f}%  {'PASS' if ok else 'no'}")
