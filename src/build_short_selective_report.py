"""Full report: apply SHORT-SELECTIVITY (only short confirmed downtrends) to the constrained
trend-aligned model. Compares base vs short-selective across aligned-caps 3.25/3.5/4.0, with full
metrics, win rates (overall/long/short), long/short P&L, 3-window drops, and constraint check.
Constraints: in&out>200, 0 liq, maxDD<=55%, $@50bp>=$1M, Calmar>1, Sharpe>1.
"""
import os
import numpy as np, pandas as pd
import stable_combo as sc
import live_engine as le
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
ANN = 365
OUT = r"C:\Users\user\OneDrive\Desktop\New setup for BTC\BTC_short_selective_report.xlsx"
df, reg0, memb = sc.prep()
reg, emap, exp_raw = le.ensemble_ctx(df, memb)
expf = np.where(np.abs(exp_raw) >= 0.4, exp_raw, 0.0)
close = df["close"].to_numpy(); low = df["low"].to_numpy(); high = df["high"].to_numpy()
sma200 = df["SMA200"].to_numpy(); n = len(df); i0 = 260; dstr = df["Date"].tolist(); dates = pd.to_datetime(df["Date"])
rv = pd.Series(close).pct_change().rolling(20).std().to_numpy() * np.sqrt(ANN)
regA = np.array(reg, dtype=object); up = close > sma200
HERE = os.path.dirname(os.path.abspath(__file__))
def lmap(p, col):
    f = pd.read_csv(p); return dict(zip(f.iloc[:, 0].astype(str), f[col]))
funding = np.array([lmap(os.path.join(HERE, "..", "data", "funding.csv"), "funding_rate").get(d, np.nan) for d in dstr])
def trk(a, w=365): return pd.Series(a).rolling(w, min_periods=60).apply(lambda x: (x.iloc[-1] >= x).mean(), raw=False).to_numpy()
vr = trk(rv); fr = trk(funding)
gl = np.ones(n); gs = np.ones(n)
for i in range(n):
    if vr[i] == vr[i] and vr[i] > 0.85: gl[i] *= 0.5; gs[i] *= 0.5
    if fr[i] == fr[i]:
        if fr[i] > 0.90: gl[i] *= 0.5
        if fr[i] < 0.10: gs[i] *= 0.5
WIN = [("2017-12-16","2018-11-18"),("2021-10-20","2022-03-09"),("2025-05-22","2025-12-01")]

def build_ein(short_selective):
    e = pd.Series(expf).ewm(span=5, adjust=False).mean().to_numpy().copy()
    if short_selective:
        for i in range(n):
            if e[i] < 0 and not (regA[i] in ("STRONG_DOWN","TREND_DOWN") and close[i] < sma200[i]): e[i] = 0.0
    return e

def sim(e_in, capA, capC=3.0, vt=1.5, slip=0.005, dd_kill=0.30, band=0.15, fee=0.0005, maint=0.01):
    sgn = np.sign(e_in); aligned = ((sgn > 0) & up) | ((sgn < 0) & ~up); cap = np.where(aligned, capA, capC)
    eqv = peak = 500.0; held = 0.0; eq = np.full(n, 500.0); E = np.zeros(n); liq = 0; fuse = []
    for i in range(i0, n):
        sig = e_in[i - 1]; g = gl[i - 1] if sig > 0 else (gs[i - 1] if sig < 0 else 1.0)
        if not (rv[i - 1] == rv[i - 1] and rv[i - 1] > 0): eq[i] = eqv; E[i] = held; continue
        e = sig * g * min(cap[i - 1], vt / rv[i - 1])
        if dd_kill > 0 and eqv < peak * (1 - dd_kill): e *= 0.5
        if band > 0 and abs(e - held) < band and not (e == 0 and held != 0): e = held
        adv = (-(low[i] / close[i - 1] - 1)) if e > 0 else ((high[i] / close[i - 1] - 1) if e < 0 else 0.0)
        if e != 0 and abs(e) * max(adv, 0) >= (1 - maint):
            eqv *= 0.01; liq += 1; held = 0.0; eq[i] = eqv; E[i] = 0.0; peak = max(peak, eqv); continue
        eqv *= (1 + e * (close[i] / close[i - 1] - 1)); eqv -= eqv * abs(e - held) * (fee + slip)
        held = e; eq[i] = max(eqv, 1e-9); E[i] = e; peak = max(peak, eqv)
        if e != 0: fuse.append(abs(e) / 5.0)
    return eq, E, liq, (np.max(fuse) if fuse else 0)
def full(e_in, capA, slip=0.005):
    eq, E, liq, mf = sim(e_in, capA, slip=slip)
    s = pd.Series(eq[i0:], index=dates.iloc[i0:].values); r = s.pct_change().dropna(); yrs = len(s) / ANN
    cagr = (s.iloc[-1] / 500) ** (1 / yrs) - 1 if s.iloc[-1] > 0 else -1
    sh = r.mean() * ANN / (r.std(ddof=1) * np.sqrt(ANN)) if r.std() > 0 else float("nan")
    dd = (s / s.cummax() - 1).min(); ww = [(s.loc[a:b]/s.loc[a:b].cummax()-1).min()*100 for a,b in WIN]
    i = i0; tr = []
    while i < n:
        sg = 1 if E[i] > 0 else (-1 if E[i] < 0 else 0)
        if sg == 0: i += 1; continue
        j = i
        while j+1 < n and (1 if E[j+1]>0 else (-1 if E[j+1]<0 else 0)) == sg: j += 1
        f0 = eq[i-1] if i > 0 else 500.0; tr.append((sg, eq[j]/f0-1, eq[j]-f0)); i = j+1
    allp = np.array([t[1] for t in tr]); L = [t for t in tr if t[0] > 0]; S = [t for t in tr if t[0] < 0]
    ow = (allp > 0).mean(); lw = (np.array([t[1] for t in L]) > 0).mean() if L else 0; sw = (np.array([t[1] for t in S]) > 0).mean() if S else 0
    lpnl = sum(t[2] for t in L); spnl = sum(t[2] for t in S)
    return dict(eq=eq, s=s, final=s.iloc[-1], cagr=cagr, sh=sh, cal=(cagr/abs(dd) if dd<0 else 0), dd=dd, liq=liq, mf=mf,
                ow=ow, lw=lw, sw=sw, nL=len(L), nS=len(S), rt=len(tr), io=len(tr)*2, ww=ww, lpnl=lpnl, spnl=spnl)

CFG = [("3.25x/3 — ALL shorts", build_ein(False), 3.25),
       ("3.25x/3 — short-SELECTIVE", build_ein(True), 3.25),
       ("3.5x/3 — short-SELECTIVE", build_ein(True), 3.5),
       ("4.0x/3 — short-SELECTIVE", build_ein(True), 4.0),
       ("4.5x/3 — short-SELECTIVE", build_ein(True), 4.5)]
res = []
print(f"{'config':30s} {'$@0bp':>12s} {'$@50bp':>12s} {'Calm':>5s} {'Shrp':>5s} {'maxDD':>6s} {'Win(L/S)':>14s} {'io':>4s} {'liq':>3s} {'mFund':>6s} PASS?")
for nm, ein, capA in CFG:
    d0 = full(ein, capA, 0.0); d = full(ein, capA, 0.005)
    pas = (d['io']>200) and (d['liq']==0) and (d['dd']>=-0.55) and (d['final']>=1_000_000) and (d['cal']>1) and (d['sh']>1)
    res.append((nm, d0['final'], d, capA, pas))
    print(f"{nm:30s} ${d0['final']:>11,.0f} ${d['final']:>11,.0f} {d['cal']:>5.2f} {d['sh']:>5.2f} {d['dd']*100:>5.0f}% {d['ow']*100:>3.0f}%({d['lw']*100:.0f}/{d['sw']*100:.0f}) {d['io']:>4d} {d['liq']:>3d} {d['mf']*100:>5.0f}%  {'PASS' if pas else 'no'}")

# ===== Excel =====
HDR=Font(bold=True,color="FFFFFF"); HF=PatternFill("solid",fgColor="1F4E2E"); TIT=Font(bold=True,size=13); GRN=PatternFill("solid",fgColor="C6EFCE")
thin=Side(style="thin",color="DDDDDD"); BD=Border(thin,thin,thin,thin)
wb=Workbook(); ws=wb.active; ws.title="Report"
ws["A1"]="Short-selective on constrained trend-aligned model — full report"; ws["A1"].font=TIT
ws["A2"]="Short-selective = only short confirmed downtrends (STRONG/TREND_DOWN & price<SMA200). Targets: in&out>200, 0 liq, DD<=55%, $50bp>=$1M, Calmar>1, Sharpe>1. Green=PASS."; ws["A2"].font=Font(italic=True,color="666666")
cols=["Config","$@0bp","$@50bp","CAGR","Calmar","Sharpe","maxDD","Win% all","Win% Long","Win% Short","Long P&L $","Short P&L $","Round-trips","In&Out","Liq","max Fund%","W1","W2","W3","PASS"]
for j,c in enumerate(cols,1):
    cc=ws.cell(4,j,c); cc.font=HDR; cc.fill=HF; cc.alignment=Alignment(horizontal="center",wrap_text=True)
for k,(nm,f0,d,capA,pas) in enumerate(res):
    r=5+k; vals=[nm,round(f0),round(d['final']),d['cagr'],d['cal'],d['sh'],d['dd'],d['ow'],d['lw'],d['sw'],round(d['lpnl']),round(d['spnl']),d['rt'],d['io'],d['liq'],d['mf'],d['ww'][0]/100,d['ww'][1]/100,d['ww'][2]/100,"PASS" if pas else "no"]
    for j,v in enumerate(vals,1):
        cc=ws.cell(r,j,v); cc.border=BD
    for j in (2,3,11,12): ws.cell(r,j).number_format='#,##0'
    for j in (4,8,9,10,16,17,18,19): ws.cell(r,j).number_format='0%'
    for j in (5,6): ws.cell(r,j).number_format='0.00'
    ws.cell(r,7).number_format='0%'
    if pas:
        for j in range(1,21): ws.cell(r,j).fill=GRN
ws.column_dimensions["A"].width=30
for c in "BCDEFGHIJKLMNOPQRST": ws.column_dimensions[c].width=11

we=wb.create_sheet("Equity Curves")
plot=res
ddt=[d.strftime("%Y-%m-%d") for d in plot[0][2]['s'].index]
we.cell(1,1,"Date")
for j,(nm,f0,d,capA,pas) in enumerate(plot,2): we.cell(1,j,nm)
for i,dd in enumerate(ddt,2):
    we.cell(i,1,dd)
    for j,(nm,f0,d,capA,pas) in enumerate(plot,2): we.cell(i,j,round(float(d['s'].iloc[i-2]),2))
for cc in we[1]: cc.font=Font(bold=True)
last=len(ddt)+1
lc=LineChart(); lc.title="Short-selective variants — equity @50bp (log)"; lc.y_axis.scaling.logBase=10; lc.height=12; lc.width=26
for j in range(2,2+len(plot)): lc.add_data(Reference(we,min_col=j,min_row=1,max_row=last),titles_from_data=True)
lc.set_categories(Reference(we,min_col=1,min_row=2,max_row=last)); we.add_chart(lc, get_column_letter(len(plot)+3)+"2")
wb.save(OUT); print("\nsaved", OUT)
